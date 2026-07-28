"""
Setup one-time / idempotente da planilha de metas (.xlsx in-place via Drive API).

CONSOLIDA tudo de cada cidade numa aba so ("Metas [ CIDADE ]"):
 - tabela semanal (Semana | Periodo | Meta | Realizado | Gap | detalhamento por tier) + grafico
 - bloco "Meta por tier" (Basico/Premium/Total: Meta | Realizado | Gap) com CF, em TODAS as cidades
 - secao "Metas Gratuitas" (Inicio | Meta | Realizado | Gap | Observacao) com CF
 - esconde as abas "Metas Gratuitas [ CIDADE ]" antigas (nao deleta -> reversivel)

DADOS: corrige Periodo BH, preenche Realizado semanal (pagas) e gratuitas a partir do raw.
VISUAL: dedup coluna duplicada, esconde Acumulado, header laranja, Calibri, #,##0, zebra, freeze.

Roda LOCAL. O cron horario (sync.py) mantem o Realizado depois. Idempotente (pode re-rodar).

Uso:  python setup_metas_xlsx.py
"""
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import sync  # noqa: E402

import gspread  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # noqa: E402
from openpyxl.formatting.rule import FormulaRule  # noqa: E402
from openpyxl.formatting.formatting import ConditionalFormattingList  # noqa: E402
# add_evolucao_chart e _last_semana_row vivem em sync.py (compartilhados com o cron auto-cura).

DASHBOARD_ID = "1KfTWNTDoWUok-yn_gGlZaJk_lhPmq9RXOh793mZomFA"

# Paleta Brada
LARANJA = "C55A11"
LARANJA_CLARO = "FBE5D6"
VERMELHO = "EA9999"
VERDE = "C6EFCE"
CINZA_TXT = "333333"
CINZA_BORDA = "D9D9D9"

HEADER_FILL = PatternFill("solid", fgColor="FF" + LARANJA)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=11, color="FF" + CINZA_TXT)
BOLD_FONT = Font(name="Calibri", bold=True, color="FF" + CINZA_TXT, size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="FF" + LARANJA_CLARO)
NO_FILL = PatternFill(fill_type=None)
_THIN = Side(style="thin", color=CINZA_BORDA)
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
NUM_COLS = set("CDEFGHIJK")

# (sigla, aba pagas original, aba gratuitas original, tem grafico)
CIDADES = {
    86595: ("BSB", "Metas Pagas [ BSB ]", "Metas Gratuitas [ BSB ]", False),
    86781: ("BH", "Metas Pagas [ BH ]", "Metas Gratuitas [ BH ]", True),
    87008: ("SSA", "Metas Pagas [ SSA ]", "Metas Gratuitas [ SSA ]", True),
}
# Metas por tier (Bruno tem so as do SSA, do print). BH/BSB ficam em branco p/ a Tamyris.
METAS_TIER = {
    "SSA": {"Básico": 1200, "Premium": 200, "Total": 1400},
    "BH": {},
    "BSB": {},
}
# Semanas a GARANTIR na tabela de cada cidade (anexadas se faltarem, idempotente). SSA: inscricoes
# vao ate 09/08, entao estende a regua segunda-a-segunda ate a semana que cobre 09/08 (06/08-13/08).
SEMANAS_EXTRA = {
    "SSA": [
        ("Semana 7", "02/07 - 09/07"),
        ("Semana 8", "09/07 - 16/07"),
        ("Semana 9", "16/07 - 23/07"),
        ("Semana 10", "23/07 - 30/07"),
        ("Semana 11", "30/07 - 06/08"),
        ("Semana 12", "06/08 - 13/08"),
    ],
}
# Posicoes dos blocos abaixo da tabela sao DINAMICAS (sync._metas_layout, pelo nº de semanas).


def reconstruir_participantes():
    gc = gspread.authorize(sync.get_credentials())
    sh = gc.open_by_key(DASHBOARD_ID)

    def parts(tab):
        rows = sh.worksheet(tab).get_all_values()[1:]
        return [{"categoria": r[1], "status": r[4], "valor": r[6], "dataPedido": r[7]}
                for r in rows if any(c.strip() for c in r)]

    return {86595: parts("raw_inscritos_brasilia"),
            86781: parts("raw_inscritos_bh"),
            87008: parts("raw_inscritos_ssa")}


def find_city_tab(wb, sigla):
    """Acha a aba da cidade pelo nome novo (consolidado) ou pelo antigo (idempotente)."""
    for cand in (f"Metas [ {sigla} ]", f"Metas Pagas [ {sigla} ]"):
        ws = sync._find_ws(wb, cand)
        if ws:
            return ws
    return None


def dedup_total_pago(ws):
    hmap = sync._header_map_xlsx(ws)
    cL = hmap.get(sync._norm_header("Real. Total Pago"))
    if not cL:
        return
    for r in range(1, ws.max_row + 1):
        ws.cell(r, cL).value = None
    ws.column_dimensions[get_column_letter(cL)].hidden = True


def simplify_meta_acumulado(ws):
    """Renomeia 'Meta Vendas Pagas'->'Meta' e esconde SO a coluna 'Acumulado'."""
    hmap = sync._header_map_xlsx(ws)
    cM = hmap.get(sync._norm_header("Meta Vendas Pagas"))
    if cM:
        ws.cell(1, cM).value = "Meta"
    cA = hmap.get(sync._norm_header("Acumulado"))
    if cA:
        # a dimensao de 'Acumulado' as vezes cobre uma FAIXA (D:E) -> colapsa p/ so a coluna dela
        cd = ws.column_dimensions[get_column_letter(cA)]
        cd.min = cA
        cd.max = cA
        cd.hidden = True
        nxt = ws.column_dimensions[get_column_letter(cA + 1)]
        nxt.min = cA + 1
        nxt.max = cA + 1
        nxt.hidden = False


def style_pagas(ws, last):
    widths = {"A": 12, "B": 16, "C": 18, "D": 12, "E": 12, "F": 10,
              "G": 13, "H": 13, "I": 11, "J": 9, "K": 13, "M": 3}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for c in range(1, 12):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r in range(2, last + 1):
        zebra = (r % 2 == 0)
        for c in range(1, 12):
            cell = ws.cell(r, c)
            col = get_column_letter(c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if col in NUM_COLS:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
            if col in ("E", "F"):
                cell.fill = NO_FILL
            elif zebra:
                cell.fill = ZEBRA_FILL
            else:
                cell.fill = NO_FILL


def _cf_meta(ws, rng, meta_anchor, real_anchor):
    """CF verde/vermelho: verde se Realizado>=Meta, vermelho se <, ignora vazio/Meta vazia."""
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f"AND(ISNUMBER({meta_anchor}),ISNUMBER({real_anchor}),{real_anchor}>={meta_anchor})"],
        fill=PatternFill("solid", fgColor="FF" + VERDE)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f"AND(ISNUMBER({meta_anchor}),ISNUMBER({real_anchor}),{real_anchor}<{meta_anchor})"],
        fill=PatternFill("solid", fgColor="FF" + VERMELHO)))


def apply_metas_cf(ws, last):
    """ZERA a CF da aba e aplica a CF semanal (Realizado E vs Meta C). Chamar ANTES de bloco/gratuitas."""
    ws.conditional_formatting = ConditionalFormattingList()
    _cf_meta(ws, f"E2:F{last}", "$C2", "$E2")


def ensure_semanas_extra(ws, sigla):
    """Garante (idempotente) as semanas de SEMANAS_EXTRA na tabela da cidade, anexadas APOS a ultima
    Semana existente. So escreve Semana (A) + Periodo (B) — a Meta e da Tamyris, Realizado e do cron
    (semanas futuras ficam em branco ate chegarem). CHAMAR depois de limpar os blocos de baixo."""
    extras = SEMANAS_EXTRA.get(sigla, [])
    if not extras:
        return
    h = sync._header_map_xlsx(ws)
    cs, cp = h.get(sync._norm_header("Semana")), h.get(sync._norm_header("Período"))
    existentes = {str(ws.cell(r, cs).value or "").strip().lower()
                  for r in range(2, ws.max_row + 1)
                  if str(ws.cell(r, cs).value or "").strip().lower().startswith("semana")}
    r = sync._last_semana_row(ws) + 1
    for label, periodo in extras:
        if label.lower() in existentes:
            continue
        ws.cell(r, cs).value = label
        ws.cell(r, cp).value = periodo
        r += 1


def clear_below_weeks(ws, last):
    """Limpa A-K (valor + estilo) abaixo da tabela semanal (blocos antigos de tier/gratuitas) p/
    reposicionar sem deixar lixo. Preserva as linhas das semanas (2..last, dados da Tamyris) e o
    painel (cols N:T). Remove charts (recriados depois pelo add_evolucao_chart)."""
    nb = Border()
    for r in range(last + 1, last + 35):
        for c in range(1, 12):  # A..K
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = NO_FILL
            cell.border = nb
    ws._charts = []


def build_tier_block(ws, metas, b):
    """Bloco 'Meta por tier' (Basico/Premium/Total): Meta | Realizado(=SUM) | Gap, com CF. `b` = linha
    do cabecalho (dinamica, via sync._metas_layout)."""
    # Cabecalhos: A=titulo, B=Meta, C=Realizado, E=Gap. O Gap fica na E (NAO na D): a coluna D
    # ('Acumulado') vive ESCONDIDA, entao um Gap na D sumiria junto. Pula a D.
    for col, h in [(1, "Meta por tier"), (2, "Meta"), (3, "Realizado"), (5, "Gap")]:
        cell = ws.cell(b, col)
        cell.value = h
        cell.font = BOLD_FONT
        cell.fill = ZEBRA_FILL
        cell.border = BORDER
    ws.cell(b, 4).value = None  # limpa header 'Gap' que setups antigos deixaram na D (escondida)
    linhas = [("Básico (R$99)", "Básico", "G"),
              ("Premium (R$159)", "Premium", "H"),
              ("Total Pago", "Total", "E")]
    for j, (label, key, sumcol) in enumerate(linhas):
        r = b + 1 + j
        # Total usa SUMIF nas linhas 'Semana*' p/ NAO somar o proprio Gap, que agora mora na coluna E.
        real = '=SUMIF($A:$A,"Semana*",$E:$E)' if sumcol == "E" else f"=SUM({sumcol}:{sumcol})"
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = metas.get(key)                       # B Meta (None = Tamyris preenche)
        ws.cell(r, 3).value = real                                 # C Realizado
        ws.cell(r, 4).value = None                                 # limpa Gap velho na D (escondida)
        ws.cell(r, 5).value = f'=IF($B{r}="","",$B{r}-$C{r})'      # E Gap (sem meta -> branco)
        for c in (1, 2, 3, 5):
            ws.cell(r, c).font = BODY_FONT
            ws.cell(r, c).border = BORDER
            if c >= 2:
                ws.cell(r, c).number_format = "#,##0"
    _cf_meta(ws, f"C{b+1}:E{b+3}", f"$B{b+1}", f"$C{b+1}")


def build_gratuitas_section(ws, src_ws, s):
    """Secao 'Metas Gratuitas' (Inicio | Meta | Realizado | Gap | Observacao) + CF. Copia dados do src.
    `s` = linha do titulo (dinamica, via sync._metas_layout)."""
    ws.cell(s, 1).value = "Metas Gratuitas"
    ws.cell(s, 1).font = BOLD_FONT
    for i, h in enumerate(["Início Monitoramento", "Meta Gratuitas", "Realizado", "Gap", "Observação"]):
        cell = ws.cell(s + 1, i + 1)
        cell.value = h
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    inicio = meta = obs = None
    if src_ws is not None and src_ws.max_row >= 2:
        sh = sync._header_map_xlsx(src_ws)
        ci = sh.get(sync._norm_header("Início Monitoramento"))
        cm = sh.get(sync._norm_header("Meta Gratuitas"))
        co = sh.get(sync._norm_header("Observação"))
        inicio = src_ws.cell(2, ci).value if ci else None
        meta = src_ws.cell(2, cm).value if cm else None
        obs = src_ws.cell(2, co).value if co else None
    d = s + 2
    ws.cell(d, 1).value = inicio
    ws.cell(d, 2).value = meta
    ws.cell(d, 4).value = f"=B{d}-C{d}"   # Gap = Meta - Realizado (Realizado e preenchido pelo cron)
    ws.cell(d, 5).value = obs
    for c in range(1, 6):
        ws.cell(d, c).font = BODY_FONT
        ws.cell(d, c).border = BORDER
    ws.cell(d, 1).number_format = "DD/MM/YYYY"
    for c in (2, 3, 4):
        ws.cell(d, c).number_format = "#,##0"
    _cf_meta(ws, f"C{d}:D{d}", f"$B{d}", f"$C{d}")


PANEL_COL0 = 14  # coluna N (inicio do painel)


def _safe_merge(ws, rng):
    """Merge idempotente: desfaz o merge identico antes de remergir (re-rodar o setup nao duplica)."""
    if rng in [str(m) for m in ws.merged_cells.ranges]:
        ws.unmerge_cells(rng)
    ws.merge_cells(rng)


def build_painel(ws, sigla, last, metas, tier_row, grat_row):
    """Painel de storytelling DURAVEL (cols N:T, linhas 2-12): KPIs por formula. Tudo sobrevive a
    edicao humana no Google (so o grafico-objeto morre -> auto-cura no cron). Identico nas 3 cidades.
    `tier_row`/`grat_row` (de sync._metas_layout) sao onde os blocos REFERENCIADOS vivem na tabela."""
    N, O, P, Q, R, T = 14, 15, 16, 17, 18, 20  # colunas do painel
    tier_basico, tier_premium, tier_total = tier_row + 1, tier_row + 2, tier_row + 3
    grat_data = grat_row + 2

    # larguras
    for col, w in {"N": 14, "O": 11, "P": 9, "Q": 9, "R": 9, "S": 9, "T": 9}.items():
        ws.column_dimensions[col].width = w

    # limpa visuais de versoes antigas que viviam em R:T (sparkline em R3:T6, barras em R9:T12)
    for rng in ("R3:T6", "R9:T9", "R10:T10", "R11:T11", "R12:T12"):
        if rng in [str(m) for m in ws.merged_cells.ranges]:
            ws.unmerge_cells(rng)
    for rr in (3, 9, 10, 11, 12):
        ws.cell(rr, R).value = None

    def _cell(r, c, val, font=BODY_FONT, fmt=None, fill=None, align=None):
        cell = ws.cell(r, c)
        cell.value = val
        cell.font = font
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = Alignment(horizontal=align, vertical="center")
        return cell

    # titulo
    _safe_merge(ws, "N2:T2")
    _cell(2, N, f"Painel {sigla}", font=HEADER_FONT, fill=HEADER_FILL, align="center")
    for c in range(N, T + 1):
        ws.cell(2, c).fill = HEADER_FILL
        ws.cell(2, c).border = BORDER

    # KPIs headline (N=rotulo, O=valor)
    _cell(3, N, "Realizado", font=BOLD_FONT)
    _cell(3, O, '=SUMIF($A:$A,"Semana*",$E:$E)', font=BOLD_FONT, fmt="#,##0", align="center")
    _cell(4, N, "Meta total")
    _cell(4, O, f'=IF($B${tier_total}="","",$B${tier_total})', fmt="#,##0", align="center")
    _cell(5, N, "% atingido")
    _cell(5, O, '=IF(OR($O$4="",$O$4=0),"",$O$3/$O$4)', fmt="0%", align="center")
    _cell(6, N, "Gap")
    _cell(6, O, '=IF($O$4="","",$O$4-$O$3)', fmt="#,##0", align="center")

    # (o gráfico de evolução fica no LineChart grande, ancorado em A20 — nao em sparkline aqui)

    # progresso por tier
    _safe_merge(ws, "N8:T8")
    _cell(8, N, "Progresso por tier", font=BOLD_FONT, fill=ZEBRA_FILL)
    for c in range(N, T + 1):
        ws.cell(8, c).fill = ZEBRA_FILL
        ws.cell(8, c).border = BORDER
    # rotulo + linha-fonte (dinamica) no bloco de tier/gratuitas
    for i, (nome, src) in enumerate([("Básico", tier_basico), ("Premium", tier_premium),
                                     ("Total", tier_total), ("Gratuitas", grat_data)]):
        r = 9 + i
        _cell(r, N, nome)
        _cell(r, O, f"=$C${src}", fmt="#,##0", align="center")                         # Realizado
        _cell(r, P, f'=IF($B${src}="","—",$B${src})', fmt="#,##0", align="center")      # Meta
        _cell(r, Q, f'=IF(OR($B${src}="",$B${src}=0),"",$C${src}/$B${src})',            # % atingido
              fmt="0%", align="center")


def main():
    drive = sync._drive_service()
    wb = openpyxl.load_workbook(io.BytesIO(sync._download_xlsx(drive, sync.METAS_SPREADSHEET_ID)))
    ppc = reconstruir_participantes()

    for eid, (sigla, pagas_tab, grat_tab, com_grafico) in CIDADES.items():
        ws = find_city_tab(wb, sigla)
        if ws is None:
            print(f"  aba da cidade {sigla} nao encontrada — pulando")
            continue
        src_g = sync._find_ws(wb, grat_tab)

        # 1) BH: corrige Periodo duplicado da Semana 6
        if sigla == "BH":
            h = sync._header_map_xlsx(ws)
            cs, cp = h.get(sync._norm_header("Semana")), h.get(sync._norm_header("Período"))
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, cs).value or "").strip().lower() == "semana 6":
                    ws.cell(r, cp).value = "11/06 - 18/06"

        # 2) ESTRUTURA: limpa blocos antigos abaixo da tabela -> anexa semanas extras (SSA) -> dados
        clear_below_weeks(ws, sync._last_semana_row(ws))
        ensure_semanas_extra(ws, sigla)
        sync.write_metas_pagas_xlsx(ws, ppc.get(eid, []), pagas_tab)  # preenche todas (futuras = branco)
        dedup_total_pago(ws)
        simplify_meta_acumulado(ws)
        last = sync._last_semana_row(ws)
        tier_row, grat_row, chart_anchor = sync._metas_layout(last)  # posicoes DINAMICAS pelo nº de semanas

        # 3) visual nas posicoes dinamicas (CF zera primeiro; bloco/gratuitas/painel add CF deles DEPOIS)
        style_pagas(ws, last)
        apply_metas_cf(ws, last)
        build_tier_block(ws, METAS_TIER.get(sigla, {}), tier_row)
        build_gratuitas_section(ws, src_g, grat_row)
        sync.write_metas_gratuitas_xlsx(ws, ppc.get(eid, []), pagas_tab)  # preenche Realizado gratuitas
        build_painel(ws, sigla, last, METAS_TIER.get(sigla, {}), tier_row, grat_row)
        if com_grafico:
            sync.add_evolucao_chart(ws, last, sigla, anchor=chart_anchor)  # grafico grande (auto-cura no cron)

        # 4) renomeia a aba (consolidada) e esconde a antiga de gratuitas
        ws.title = pagas_tab.replace("Pagas ", "").replace("[", "").replace("]", "")  # "Metas  CIDADE "
        if src_g is not None:
            src_g.sheet_state = "hidden"
        print(f"  consolidado: {ws.title!r} (last={last}, tier@{tier_row}, grat@{grat_row}, chart@{chart_anchor}, grafico={com_grafico})")

    out = io.BytesIO()
    wb.save(out)
    sync._upload_xlsx(drive, sync.METAS_SPREADSHEET_ID, out.getvalue())
    print("Consolidacao + visual aplicados (upload OK).")


if __name__ == "__main__":
    main()
