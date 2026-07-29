"""
Testes unitarios (sem rede) da logica de metas em sync.py.
Rodar: python test_metas.py   (de dentro de github-actions/)
"""
from datetime import date

import sync

_fail = 0


def chk(name, got, exp):
    global _fail
    ok = got == exp
    print(("OK  " if ok else "FAIL") + f" {name}: got={got!r} exp={exp!r}")
    if not ok:
        _fail += 1


def P(cat, valor, status="Pago", data="01/06/2026 10:00"):
    return {"categoria": cat, "valor": valor, "status": status, "dataPedido": data}


# --- parse_valor ---
chk("valor_str", sync.parse_valor("99.00"), 99.0)
chk("valor_virgula", sync.parse_valor("109,90"), 109.9)
chk("valor_num", sync.parse_valor(159), 159.0)
chk("valor_none", sync.parse_valor(None), 0.0)
chk("valor_vazio", sync.parse_valor(""), 0.0)
chk("valor_zero", sync.parse_valor("0,00"), 0.0)

# --- parse_data_pedido ---
chk("data_ok", sync.parse_data_pedido("13/05/2026 11:02"), date(2026, 5, 13))
chk("data_sec", sync.parse_data_pedido("13/05/2026 11:02:33"), date(2026, 5, 13))
chk("data_ruim", sync.parse_data_pedido(""), None)
chk("data_lixo", sync.parse_data_pedido("sem data"), None)

# --- parse_periodo_fim / parse_inicio ---
chk("periodo", sync.parse_periodo_fim("14/05 - 21/05"), date(2026, 5, 21))
chk("periodo_endash", sync.parse_periodo_fim("14/05 – 21/05"), date(2026, 5, 21))
chk("periodo_ruim", sync.parse_periodo_fim("xxx"), None)
chk("periodo_vazio", sync.parse_periodo_fim(""), None)
chk("inicio", sync.parse_inicio("01/06"), date(2026, 6, 1))
chk("inicio_ano", sync.parse_inicio("01/06/2026"), date(2026, 6, 1))

# --- classificacao (categorias reais das 3 cidades) ---
chk("base_basico", sync.base_tier(P("Kit Vai Bem", "99.00")), "Básico")
chk("base_oculto", sync.base_tier(P("Kit Vai Bem - Oculto", "99.00")), "Básico")
chk("base_premium", sync.base_tier(P("KIT PREMIUM", "159.00")), "Premium")
chk("base_premium_combo", sync.base_tier(P("KIT PREMIUM - COMBO DIA DAS MAES", "222.60")), "Premium")
chk("is_combo", sync.is_combo(P("Kit VAI BEM - COMBO DIA DOS NAMORADOS", "109,90")), True)
chk("is_combo_nao", sync.is_combo(P("Kit Vai Bem", "99.00")), False)
chk("is_pcd", sync.is_pcd(P("KIT PREMIUM - PCD", "79,50")), True)
chk("free_valor0", sync.is_free(P("Kit VAI BEM - COMBO DIA DOS NAMORADOS", "0,00")), True)
chk("free_cortesia", sync.is_free(P("Kit Vai Bem - Oculto", "99.00", status="Cortesia")), True)
chk("free_nao", sync.is_free(P("Kit Vai Bem", "99.00")), False)

# --- tier_counts_cumulative ---
amostra = [
    P("Kit Vai Bem", "99.00", data="10/05/2026 10:00"),                                   # basico pago
    P("KIT PREMIUM", "159.00", data="12/05/2026 10:00"),                                  # premium pago
    P("Kit VAI BEM - COMBO DIA DOS NAMORADOS", "109,90", data="12/05/2026 11:00"),        # combo pago (base basico)
    P("Kit VAI BEM - COMBO DIA DOS NAMORADOS", "0,00", data="12/05/2026 11:00"),          # combo gratis
    P("KIT PREMIUM - PCD", "79,50", data="20/05/2026 10:00"),                             # premium+pcd pago (20/05)
    P("Kit Vai Bem - Oculto", "99.00", status="Cortesia", data="09/05/2026 10:00"),       # cortesia -> gratuito
]
c = sync.tier_counts_cumulative(amostra, date(2026, 5, 14))  # exclui o de 20/05
chk("cum_total_pago", c["Total Pago"], 3)
chk("cum_basico", c["Básico"], 2)
chk("cum_premium", c["Premium"], 1)
chk("cum_combo", c["Combo"], 1)
chk("cum_pcd", c["PCD"], 0)
chk("cum_gratuito", c["Gratuito"], 2)
chk("cum_total_eq_base", c["Básico"] + c["Premium"], c["Total Pago"])

c2 = sync.tier_counts_cumulative(amostra, date(2026, 5, 31))  # inclui o de 20/05
chk("cum2_total_pago", c2["Total Pago"], 4)
chk("cum2_premium", c2["Premium"], 2)
chk("cum2_pcd", c2["PCD"], 1)

# --- gratuito_count_since ---
chk("grat_since_10mai", sync.gratuito_count_since(amostra, date(2026, 5, 10)), 1)  # cortesia 09/05 fora
chk("grat_since_none", sync.gratuito_count_since(amostra, None), 2)

# --- _to_int (Acumulado/Meta) ---
chk("toint_str", sync._to_int("1200"), 1200)
chk("toint_milhar", sync._to_int("1.200"), 1200)
chk("toint_num", sync._to_int(510), 510)
chk("toint_vazio", sync._to_int(""), None)

# --- _semana_futura (fill-as-time) ---
HOJE = date(2026, 6, 21)
chk("futura_sim", sync._semana_futura("25/06 - 02/07", HOJE), True)     # inicio 25/06 > 21/06
chk("futura_borda", sync._semana_futura("21/06 - 28/06", HOJE), False)  # comeca hoje -> mostra
chk("futura_atual", sync._semana_futura("18/06 - 25/06", HOJE), False)  # semana corrente
chk("futura_passada", sync._semana_futura("14/05 - 21/05", HOJE), False)
chk("futura_endash", sync._semana_futura("25/06 – 02/07", HOJE), True)  # en-dash
chk("futura_lixo", sync._semana_futura("xxx", HOJE), False)             # nao-parseavel -> nao esconde

# --- painel + grafico (storytelling duravel; nada de rede) ---
import io as _io         # noqa: E402
import os as _os         # noqa: E402
import sys as _sys       # noqa: E402
import openpyxl as _oxl  # noqa: E402
_sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
import setup_metas_xlsx as _S  # noqa: E402


def _fake_tab():
    wb = _oxl.Workbook()
    ws = wb.active
    ws.title = "Metas BH"
    hdr = ["Semana", "Período", "Meta", "Acumulado", "Realizado", "Gap",
           "Real. Básico", "Real. Premium", "Real. Combo", "Real. PCD", "Real. Gratuito"]
    for c, h in enumerate(hdr, start=1):
        ws.cell(1, c).value = h
    semanas = [("Semana 0", "30/04 - 07/05"), ("Semana 1", "07/05 - 14/05"),
               ("Semana 2", "14/05 - 21/05"), ("Semana 3", "21/05 - 28/05"),
               ("Semana 4", "28/05 - 04/06"), ("Semana 5", "04/06 - 11/06"),
               ("Semana 6", "11/06 - 18/06")]
    for i, (s, p) in enumerate(semanas):
        r = 2 + i
        ws.cell(r, 1).value = s
        ws.cell(r, 2).value = p
        ws.cell(r, 3).value = 100 + i * 10   # Meta semanal
        ws.cell(r, 5).value = 50 + i * 5     # Realizado semanal
        ws.cell(r, 7).value = 40 + i * 5     # Real. Basico
        ws.cell(r, 8).value = 10             # Real. Premium
    # so o que o painel referencia do bloco de tier (11-13) e gratuitas (17)
    ws.cell(11, 2).value, ws.cell(11, 3).value = 1200, 672
    ws.cell(12, 2).value, ws.cell(12, 3).value = 200, 160
    ws.cell(13, 2).value, ws.cell(13, 3).value = 1400, 832
    ws.cell(17, 2).value, ws.cell(17, 3).value = 300, 99
    return wb, ws


# layout dinamico (posicoes pelo nº de semanas)
chk("layout_bh", sync._metas_layout(8), (10, 15, "A19"))
chk("layout_ssa_estendido", sync._metas_layout(14), (16, 21, "A25"))

_wb, _ws = _fake_tab()
_before_AK = [[_ws.cell(r, c).value for c in range(1, 12)] for r in range(2, 9)]
# fake tab tem tier@10-13 e grat@15-17 -> tier_row=10, grat_row=15
_S.build_painel(_ws, "BH", 8, {"Básico": 1200, "Premium": 200, "Total": 1400}, 10, 15)
sync.add_evolucao_chart(_ws, 8, "BH")

chk("painel_titulo", _ws.cell(2, 14).value, "Painel BH")
chk("painel_realizado_sum", _ws.cell(3, 15).value, '=SUMIF($A:$A,"Semana*",$E:$E)')
chk("painel_tier_basico_real", _ws.cell(9, 15).value, "=$C$11")
chk("painel_tier_basico_pct", str(_ws.cell(9, 17).value).startswith("=IF(OR($B$11"), True)
chk("painel_sem_sparkline", _ws.cell(3, 18).value, None)   # R3 vazio (sparkline removido)
chk("painel_AK_intacto", [[_ws.cell(r, c).value for c in range(1, 12)] for r in range(2, 9)], _before_AK)
chk("painel_chart_1", len(_ws._charts), 1)
chk("painel_chart_anchor_A20", _ws._charts[0].anchor, "A20")  # em memoria e a string; vira A20 (col0,row19) ao salvar

# sobrevive ao round-trip do openpyxl (== o que o cron faz: load -> save)
_buf = _io.BytesIO()
_wb.save(_buf)
_ws2 = _oxl.load_workbook(_io.BytesIO(_buf.getvalue())).worksheets[0]
chk("painel_rt_titulo", _ws2.cell(2, 14).value, "Painel BH")
chk("painel_rt_chart_sem_dup", len(_ws2._charts), 1)

# --- bloco de tier: Gap na coluna E (NAO na D escondida) ---
_wbt = _oxl.Workbook()
_wst = _wbt.active
_wst.cell(2, 1).value = "Semana 1"   # 1 linha 'Semana*' pro SUMIF
_wst.cell(2, 5).value = 100
_S.build_tier_block(_wst, {"Básico": 1200, "Premium": 200, "Total": 1400}, 10)
chk("tier_header_gap_em_E", _wst.cell(10, 5).value, "Gap")          # E10
chk("tier_header_D_vazio", _wst.cell(10, 4).value, None)            # D10 sem 'Gap'
chk("tier_gap_total_em_E", str(_wst.cell(13, 5).value).startswith('=IF($B13'), True)  # E13 Gap guardado
chk("tier_gap_D_vazio", _wst.cell(13, 4).value, None)              # D13 sem Gap
chk("tier_total_realizado_sumif", _wst.cell(13, 3).value, '=SUMIF($A:$A,"Semana*",$E:$E)')

# --- write_metas_native: _build_metas_writes (mapeamento/branco/gratuitas, sem rede) ---
HOJE_N = date(2026, 6, 22)
_grid = [
    ["Semana", "Período", "Meta", "Acumulado", "Realizado", "Gap",
     "Real. Básico", "Real. Premium", "Real. Combo", "Real. PCD", "Real. Gratuito"],
    ["Semana 0", "30/04 - 07/05"],   # passada, 0 inscritos ate 07/05
    ["Semana 1", "07/05 - 14/05"],   # passada, 2 pagos
    ["Semana 2", "25/06 - 02/07"],   # FUTURA (inicio 25/06 > 22/06) -> clear
    [],
    ["Meta por tier", "Meta", "Realizado"],   # NAO e linha 'Semana' -> ignora
    ["Básico (R$99)"],
    ["Metas Gratuitas"],
    ["Início Monitoramento", "Meta Gratuitas", "Realizado", "Gap", "Observação"],
    ["01/05", 300, "", "", "obs"],   # gratuitas: dados na linha 10
]
_parts = [P("Kit Vai Bem", "99.00", data="10/05/2026 10:00"),                          # basico, Semana 1
          P("KIT PREMIUM", "159.00", data="12/05/2026 10:00"),                         # premium, Semana 1
          P("Kit Vai Bem - Oculto", "99.00", status="Cortesia", data="03/05/2026 10:00")]  # gratuito desde 01/05
_u, _c, _log = sync._build_metas_writes(_grid, _parts, "TEST", HOJE_N)
_um = {d["range"]: d["values"][0] for d in _u}
_starts = [r.split(":")[0] for r in _um]
chk("nat_clear_futura", "E4:K4" in _c, True)               # Semana 2 (linha 4) futura -> clear
chk("nat_sem0_zero", _um["E2:K2"][0], 0)                   # Semana 0: 0 pagos -> escreve 0 (nao branco)
chk("nat_gap_formula", _um["E3:K3"][1], "=C3-E3")          # F = formula sem separador (imune a locale)
chk("nat_sem1_total", _um["E3:K3"][0], 2)                  # Semana 1: 2 pagos ate 14/05
chk("nat_sem1_basico", _um["E3:K3"][2], 1)                 # G: 1 basico
chk("nat_sem1_premium", _um["E3:K3"][3], 1)                # H: 1 premium
chk("nat_no_tierblock", any(s in ("E6", "E7") for s in _starts), False)  # bloco tier nao escrito
chk("nat_grat_real_C", _um["C10"][0], 1)                   # gratuitas Realizado em col C (header-relativo)
chk("nat_grat_gap_D", _um["D10"][0], "=B10-C10")           # gratuitas Gap em col D
chk("nat_never_AB", any(s[0] in "AB" for s in _starts), False)            # nunca escreve A/B (humano)
chk("nat_C_only_grat", all(s == "C10" for s in _starts if s[0] == "C"), True)  # C so na gratuitas

# =========================================================================
# ABAS DE META DOS PEDAIS (setup_metas_pedal.py + write_metas_native)
# =========================================================================
import setup_metas_pedal as SP  # noqa: E402

# --- regua semanal derivada (1a venda -> vespera do evento) ---
# cada config carrega a propria meta: os pedais herdam 500, Santos declara 1000
_metas_cfg = {k: c.get("meta_total", SP.META_TOTAL) for k, c in SP.CONFIGS.items()}
_reguas = {k: SP.semanas(c["inicio"], c["fim"], _metas_cfg[k]) for k, c in SP.CONFIGS.items()}
for _k, _r in _reguas.items():
    chk(f"regua_soma_meta_{_k}", sum(m for _, _, m in _r), _metas_cfg[_k])
    chk(f"regua_sem_periodo_dup_{_k}", len({p for _, p, _ in _r}), len(_r))
    # bordas encaixam em TODAS as reguas, nao so na do Road
    chk(f"regua_bordas_encaixam_{_k}",
        all(_r[i][1].split(" - ")[1] == _r[i + 1][1].split(" - ")[0] for i in range(len(_r) - 1)),
        True)
chk("pedal_n_road", len(_reguas["pedalx_road"]), 5)
chk("pedal_n_manaus", len(_reguas["pedalx_manaus"]), 7)
chk("pedal_n_canastra", len(_reguas["pedalx_canastra"]), 12)
# Santos: 29/07 -> 14/09 (signUpDeadLine), meta 1000 em 7 semanas
chk("santos_n_semanas", len(_reguas["circuito_santos"]), 7)
chk("santos_meta_nao_herdou_500", _metas_cfg["circuito_santos"], 1000)
chk("santos_primeira_semana", _reguas["circuito_santos"][0][1], "29/07 - 05/08")
chk("santos_ultima_semana", _reguas["circuito_santos"][-1][1], "09/09 - 14/09")
# os pedais NAO podem ter herdado a meta do Santos (regressao da parametrizacao)
chk("pedais_seguem_em_500",
    {_metas_cfg[k] for k in ("pedalx_road", "pedalx_manaus", "pedalx_canastra")}, {500})
# ultima semana pode ser curta e termina no ULTIMO DIA DE VENDA lido no site,
# nao na vespera do evento (Road: vende ate 23/08, evento so em 30/08)
chk("pedal_ultima_semana_road", _reguas["pedalx_road"][-1][1], "18/08 - 23/08")
# janela invalida deve estourar, nao gerar regua vazia
try:
    SP.semanas(date(2026, 8, 30), date(2026, 8, 30))
    chk("pedal_janela_invalida", "sem erro", "ValueError")
except ValueError:
    chk("pedal_janela_invalida", "ValueError", "ValueError")

# --- layout gerado: zebra, acumulado, bloco de tier ---
_g, _m = SP.build_layout(SP.CONFIGS["pedalx_road"])


def _bg(r, letra):
    cd = _g[r - 1][SP.col_idx(letra)]
    return (cd.get("userEnteredFormat") or {}).get("backgroundColor")


def _v(r, letra):
    uev = _g[r - 1][SP.col_idx(letra)].get("userEnteredValue") or {}
    return uev.get("formulaValue", uev.get("stringValue", uev.get("numberValue")))


# E e F NUNCA levam zebra: a formatacao condicional e dona dessas duas colunas (igual SSA)
chk("pedal_zebra_par_em_C", _bg(2, "C"), SP.ZEBRA)
chk("pedal_zebra_nunca_em_E", _bg(2, "E"), None)
chk("pedal_zebra_nunca_em_F", _bg(2, "F"), None)
chk("pedal_sem_zebra_impar", _bg(3, "C"), None)
chk("pedal_acum_primeira", _v(2, "D"), "=C2")
chk("pedal_acum_encadeado", _v(_m["wkN"], "D"), f'=D{_m["wkN"]-1}+C{_m["wkN"]}')
chk("pedal_tier_sem_meta", _v(_m["tier"], "B"), None)       # meta so na linha Total Pago
chk("pedal_total_com_meta", _v(_m["total"], "B"), 500)
chk("pedal_total_sumif_ptbr", _v(_m["total"], "C"), '=SUMIF($A:$A;"Semana*";$E:$E)')
chk("pedal_coluna_D_visivel",
    any("hiddenByUser" in str(r) for r in SP.build_requests(1, SP.CONFIGS["pedalx_road"])[0]), False)

# --- o grid do builder tem que ser legivel pelo writer do cron (as duas metades encaixam) ---
def _as_text(grid):
    """CellData -> list-of-lists de string, como ws.get_values() devolveria."""
    out = []
    for row in grid:
        linha = []
        for cd in row:
            uev = cd.get("userEnteredValue") or {}
            if "stringValue" in uev:
                linha.append(uev["stringValue"])
            elif "numberValue" in uev:
                linha.append(str(int(uev["numberValue"])))
            elif "formulaValue" in uev:
                linha.append("0")          # o Sheets devolveria o valor calculado
            else:
                linha.append("")
        out.append(linha)
    return out


_gt = _as_text(_g)
chk("pedal_schema_guard_A", sync._norm_header(_gt[0][0]), sync._norm_header("Semana"))
chk("pedal_schema_guard_E", sync._norm_header(_gt[0][4]), sync._norm_header("Realizado"))
chk("pedal_schema_guard_F", sync._norm_header(_gt[0][5]), sync._norm_header("Gap"))

# 25/07: dentro da Semana 0 (21/07-28/07). Semanas 2+ ainda sao futuras nessa data.
_HOJE_P = date(2026, 7, 25)
_pp = [P("Pedal X", "139,00", data="22/07/2026 09:00"),
       P("Pedal X", "139,00", data="24/07/2026 09:00")]
_pu, _pc, _plog = sync._build_metas_writes(_gt, _pp, "Metas Pedal Road", _HOJE_P)
_pum = {d["range"]: d["values"][0] for d in _pu}
chk("pedal_sem0_total", _pum["E2:K2"][0], 2)
chk("pedal_sem0_gap", _pum["E2:K2"][1], "=C2-E2")
# tier unico: tudo cai em Basico (G); Premium/Combo/PCD zerados
chk("pedal_sem0_basico_igual_total", _pum["E2:K2"][2], _pum["E2:K2"][0])
chk("pedal_sem0_sem_premium", _pum["E2:K2"][3:6], [0, 0, 0])
chk("pedal_semana_curta_parseia", sync.parse_periodo_fim("25/08 - 29/08"), date(2026, 8, 29))
# em 25/07 so a Semana 0 ja comecou; da Semana 1 (inicio 28/07) em diante e tudo futuro
chk("pedal_futuras_limpas", sorted(_pc),
    [f"E{r}:K{r}" for r in range(_m["wk0"] + 1, _m["wkN"] + 1)])
# gratuitas com Meta em branco: Realizado ainda e escrito e o Gap sai como formula
chk("pedal_grat_real", _pum[f'C{_m["grat"]}'][0], 0)
chk("pedal_grat_gap", _pum[f'D{_m["grat"]}'][0], f'=B{_m["grat"]}-C{_m["grat"]}')
_pstarts = [r.split(":")[0] for r in _pum]
chk("pedal_nunca_AB", any(s[0] in "AB" for s in _pstarts), False)
chk("pedal_D_so_na_gratuitas",
    all(s == f'D{_m["grat"]}' for s in _pstarts if s[0] == "D"), True)

# --- layout do Circuito Santos: rateio de gratuitas, tier proprio, ancora do grafico ---
_sg, _sm = SP.build_layout(SP.CONFIGS["circuito_santos"])


def _val(grid, r, letra):
    v = grid[r - 1][SP.col_idx(letra)].get("userEnteredValue", {})
    return v.get("stringValue", v.get("numberValue", v.get("formulaValue")))


chk("santos_tier_label", _val(_sg, _sm["tier"], "A"), "Inscrição R$70 (tier único)")
chk("santos_tier_sem_meta", _val(_sg, _sm["tier"], "B"), None)
chk("santos_total_pago_1000", _val(_sg, _sm["total"], "B"), 1000)
# a linha do cron e hdr+1 (== marcos["grat"]) e e a UNICA com meta total e Realizado
chk("santos_grat_meta_1000", _val(_sg, _sm["grat"], "B"), 1000)
chk("santos_rateio_2_linhas", _sm["rateio"], 2)
chk("santos_rateio_rotulos",
    [_val(_sg, _sm["grat"] + 1 + k, "A") for k in range(2)],
    ["Asia Shipping (garantido)", "A conquistar"])
chk("santos_rateio_metas", [_val(_sg, _sm["grat"] + 1 + k, "B") for k in range(2)], [600, 400])
chk("santos_rateio_fecha_o_total",
    sum(_val(_sg, _sm["grat"] + 1 + k, "B") for k in range(2)), _val(_sg, _sm["grat"], "B"))
# rateio NAO pode ter Realizado: a plataforma nao separa a origem da cortesia
chk("santos_rateio_sem_realizado",
    [_val(_sg, _sm["grat"] + 1 + k, "C") for k in range(2)], [None, None])
# os pedais continuam sem rateio e com o rotulo de R$139
_rg, _rm = SP.build_layout(SP.CONFIGS["pedalx_road"])
chk("pedal_sem_rateio", _rm["rateio"], 0)
chk("pedal_tier_label_intacto", _val(_rg, _rm["tier"], "A"), "Inscrição R$139 (tier único)")
chk("pedal_grat_sem_meta", _val(_rg, _rm["grat"], "B"), None)
# grafico: a folga de 2 linhas abaixo do conteudo vale em TODAS as abas, com ou sem rateio
for _k in SP.CONFIGS:
    _rq, _mk = SP.build_requests(999, SP.CONFIGS[_k])
    _anc = [r for r in _rq if "addChart" in r][0]["addChart"]["chart"]["position"]
    chk(f"grafico_folga_2_{_k}",
        _anc["overlayPosition"]["anchorCell"]["rowIndex"] - _mk["last"], 2)

# --- guarda: etapa nao sincronizada != etapa com zero inscritos ---
import gspread as _gs  # noqa: E402


class _SheetEspiao:
    """Registra que abas foram pedidas. Levanta WorksheetNotFound pra nao seguir pra rede."""

    def __init__(self):
        self.pedidas = []

    def worksheet(self, nome):
        self.pedidas.append(nome)
        raise _gs.exceptions.WorksheetNotFound(nome)


_esp = _SheetEspiao()
sync.write_metas_native(_esp, {})
chk("guarda_run_vazio_nao_toca_aba", _esp.pedidas, [])
_esp2 = _SheetEspiao()
sync.write_metas_native(_esp2, {87735: []})
chk("guarda_chave_presente_processa", _esp2.pedidas, ["Metas Pedal Road"])

# --- registro dos 7 eventos e protecao no outro builder ---
import redesign_dashboard as _rd  # noqa: E402

chk("metas_native_7_eventos", len(sync.METAS_TABS_NATIVE), 7)
chk("metas_native_cobre_pedais_e_santos",
    {87735, 87732, 87727, 87817} <= set(sync.METAS_TABS_NATIVE), True)
# O .xlsx e da Tamyris e so tem as 3 corridas; tudo que o setup_metas_pedal.py constroi
# fica fora dele. Santos entrou em 29/07 pelo mesmo motivo dos pedais.
chk("nativo_only_fora_do_xlsx",
    set(sync.METAS_TABS_NATIVE) - set(sync.METAS_TABS), {87735, 87732, 87727, 87817})
chk("builder_pedal_bate_com_nativo_only",
    sync.METAS_TABS_BUILDER_PEDAL, set(sync.METAS_TABS_NATIVE) - set(sync.METAS_TABS))
_titulos_pedal = {c["tab"] for c in SP.CONFIGS.values()}
chk("titulos_batem_com_sync", _titulos_pedal <= set(sync.METAS_TABS_NATIVE.values()), True)
chk("titulos_protegidos_no_builder", _titulos_pedal <= _rd.PROTECTED_TITLES, True)
# nenhuma aba de meta pode ser dash_tab de config do builder (senao _validate_custom recusaria)
chk("meta_nao_e_dash_tab",
    _titulos_pedal & {c["dash_tab"] for c in _rd.DASHBOARD_CONFIGS.values()}, set())

print()
if _fail:
    print(f"=== {_fail} FALHAS ===")
    raise SystemExit(1)
print("=== TODOS OS TESTES PASSARAM ===")
