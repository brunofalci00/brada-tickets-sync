"""
Build ONE-TIME das abas de metas NATIVAS no Dashboard_Inscricoes_VaiBem (Sheets API).

Transcreve FIELMENTE as abas consolidadas do .xlsx (meta_corrida_vai_bem.xlsx) para abas
nativas novas (valores + formulas + number formats + fills + fontes + bordas + merges +
colunas ocultas + freeze + larguras), e adiciona CF NATIVA + grafico NATIVO (duravel, nao
morre com edicao humana). Substitui a camada openpyxl/xlsx pela nativa.

POR QUE transcrever (e nao reconstruir): preserva os quirks manuais da Tamyris (Meta como
formula em BH C4='=170+F3', Acumulado misto, datas) e mantem o painel correto porque as
formulas absolutas ($C$17, $B$19) caem nas mesmas linhas.

LOCALE (provado por probe 22/06): esta planilha parseia formula no locale pt_BR em TODOS os
caminhos de escrita -> separador de argumento e ';' (virgula da #ERROR). As formulas do xlsx
(openpyxl) vem em virgula -> convertidas p/ ';' fora de aspas. Decimais: nenhuma no layout.

SEGURANCA: guard PROTECTED_IDS aborta se qualquer request mirar uma das 6 abas de producao.
Idempotente: se ja existir aba "Metas SSA/BH/BSB", deleta e recria limpa (sem dup de CF/grafico).
SSA/BH visiveis, BSB oculta (evento passado) e sem grafico.

Uso:  python setup_metas_native.py --dry-run    # so monta e conta os requests
      python setup_metas_native.py              # cria de verdade
"""
import argparse
import io
from datetime import date, datetime

import openpyxl
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from gspread.utils import a1_range_to_grid_range

SA = r"C:\Users\bruno\.brada-secrets\sheets-sa.json"
DEST = "1KfTWNTDoWUok-yn_gGlZaJk_lhPmq9RXOh793mZomFA"   # Dashboard nativo (destino)
SRC = "1t5xEHgT-g6k9wAWspjXKDMssX0rNYhJS"               # meta_corrida_vai_bem.xlsx (fonte)

# As 6 abas de producao que NUNCA podem ser tocadas (invariante de codigo).
PROTECTED_IDS = {73024293, 1879910525, 1573886114, 405115810, 1304144837, 757333826}

# (titulo nativo, titulo no xlsx, oculta?, tem grafico?)
CITIES = [
    ("Metas SSA", "Metas  SSA ", False, True),
    ("Metas BH",  "Metas  BH ",  False, True),
    ("Metas BSB", "Metas  BSB ", True,  False),
]

# Larguras em px (px = round(width_excel*7)+5). L fica no default; BSB oculta a L.
WIDTHS = {"A": 89, "B": 117, "C": 131, "D": 89, "E": 89, "F": 75, "G": 96, "H": 96,
          "I": 82, "J": 68, "K": 96, "M": 26, "N": 103, "O": 82, "P": 68, "Q": 68,
          "R": 68, "S": 68, "T": 68}

GRAY = "FFD9D9D9"


# ----------------------------- helpers -----------------------------

def col_idx(letter):
    i = 0
    for ch in letter:
        i = i * 26 + (ord(ch.upper()) - 64)
    return i - 1


def hexcol(argb):
    """ARGB/RGB openpyxl -> {red,green,blue} 0..1 (dropa alpha). None se invalido."""
    if not argb:
        return None
    s = str(argb)
    if len(s) == 8:
        s = s[2:]
    if len(s) != 6:
        return None
    try:
        return {"red": int(s[0:2], 16) / 255, "green": int(s[2:4], 16) / 255, "blue": int(s[4:6], 16) / 255}
    except ValueError:
        return None


def map_nf(nf):
    if not nf or nf == "General":
        return None
    if nf.strip() == "0%":
        return {"type": "PERCENT", "pattern": "0%"}
    if nf.upper() in ("DD/MM/YYYY", "DD/MM/AAAA"):
        return {"type": "DATE", "pattern": "dd/mm/yyyy"}  # token do Sheets e minusculo
    return {"type": "NUMBER", "pattern": nf}


def commas_to_semicolons(formula):
    """Troca virgula->';' SO fora de aspas (separador de argumento no locale pt_BR)."""
    out, in_str = [], False
    for ch in formula:
        if ch == '"':
            in_str = not in_str
            out.append(ch)
        elif ch == "," and not in_str:
            out.append(";")
        else:
            out.append(ch)
    return "".join(out)


def serial(dt):
    """Datetime/date -> serial Sheets (dias desde 1899-12-30)."""
    return (date(dt.year, dt.month, dt.day) - date(1899, 12, 30)).days


def extract_format(c):
    fmt = {}
    nf = map_nf(c.number_format)
    if nf:
        fmt["numberFormat"] = nf
    try:
        if c.fill and c.fill.patternType:
            col = hexcol(getattr(c.fill.fgColor, "rgb", None))
            if col:
                fmt["backgroundColor"] = col
    except Exception:
        pass
    tf = {}
    try:
        if c.font:
            if c.font.bold:
                tf["bold"] = True
            fc = getattr(c.font.color, "rgb", None) if c.font.color else None
            col = hexcol(fc) if isinstance(fc, str) else None
            if col:
                tf["foregroundColor"] = col
    except Exception:
        pass
    if tf:
        fmt["textFormat"] = tf
    try:
        al = c.alignment
        if al:
            h = {"left": "LEFT", "center": "CENTER", "right": "RIGHT"}.get(al.horizontal or "")
            if h:
                fmt["horizontalAlignment"] = h
            v = {"top": "TOP", "center": "MIDDLE", "bottom": "BOTTOM"}.get(al.vertical or "")
            if v:
                fmt["verticalAlignment"] = v
            if al.wrap_text:
                fmt["wrapStrategy"] = "WRAP"
    except Exception:
        pass
    try:
        sides = {}
        for name in ("top", "bottom", "left", "right"):
            side = getattr(c.border, name, None)
            if side and side.style:
                sides[name] = {"style": "SOLID", "color": hexcol(GRAY)}
        if sides:
            fmt["borders"] = sides
    except Exception:
        pass
    return fmt


def cell_to_celldata(c):
    cd = {}
    v = c.value
    uev = None
    if v is None or v == "":
        uev = None
    elif isinstance(v, bool):
        uev = {"boolValue": v}
    elif isinstance(v, (int, float)):
        uev = {"numberValue": v}
    elif isinstance(v, (datetime, date)):
        uev = {"numberValue": serial(v)}
    elif isinstance(v, str) and v.startswith("="):
        uev = {"formulaValue": commas_to_semicolons(v)}
    else:
        uev = {"stringValue": str(v)}
    if uev is not None:
        cd["userEnteredValue"] = uev
    fmt = extract_format(c)
    if fmt:
        cd["userEnteredFormat"] = fmt
    return cd


def used_bounds(wf, max_r=45, max_c=20):
    last_r = 1
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            if wf.cell(r, c).value not in (None, ""):
                last_r = r
                break
    return last_r


def last_semana_row(wf):
    last = 1
    for r in range(2, wf.max_row + 1):
        v = wf.cell(r, 1).value
        if v and str(v).strip().lower().startswith("semana"):
            last = r
    return last


def cf_color(rule):
    try:
        dxf = rule.dxf
        for attr in (dxf.fill.fgColor, dxf.fill.bgColor):
            rgb = getattr(attr, "rgb", None)
            if isinstance(rgb, str) and rgb not in ("00000000",):
                c = hexcol(rgb)
                if c:
                    return c
    except Exception:
        pass
    return None


def chart_request(sid, sigla, last, extra=0):
    """Grafico de evolucao semanal. `last` = ultima linha de semana (1-based).

    A ancora fica em `last + 10 + extra`, o que deixa duas linhas de folga abaixo do bloco
    padrao (8 linhas sob as semanas). `extra` existe para layouts que crescem esse bloco —
    ex.: as linhas de rateio de gratuitas do Circuito Santos. Default 0 preserva as abas
    de corrida e de pedal exatamente como estao.
    """
    def rng(c0, c1):
        return {"sheetId": sid, "startRowIndex": 0, "endRowIndex": last,
                "startColumnIndex": c0, "endColumnIndex": c1}
    return {"addChart": {"chart": {
        "spec": {"title": f"Evolução semanal {sigla}", "basicChart": {
            "chartType": "LINE", "legendPosition": "BOTTOM_LEGEND", "headerCount": 1,
            "axis": [{"position": "BOTTOM_AXIS", "title": "Semana"},
                     {"position": "LEFT_AXIS", "title": "Inscricoes"}],
            "domains": [{"domain": {"sourceRange": {"sources": [rng(0, 1)]}}}],
            "series": [
                {"series": {"sourceRange": {"sources": [rng(4, 5)]}}, "targetAxis": "LEFT_AXIS",
                 "color": hexcol("FFC55A11"), "lineStyle": {"type": "SOLID", "width": 3}},
                {"series": {"sourceRange": {"sources": [rng(2, 3)]}}, "targetAxis": "LEFT_AXIS",
                 "color": hexcol("FF999999"), "lineStyle": {"type": "MEDIUM_DASHED", "width": 2}},
            ]}},
        "position": {"overlayPosition": {
            "anchorCell": {"sheetId": sid, "rowIndex": last + 10 + extra, "columnIndex": 0},
            "offsetXPixels": 0, "offsetYPixels": 0, "widthPixels": 720, "heightPixels": 320}}
    }}}


def _collect_ids(o, acc):
    if isinstance(o, dict):
        for k, v in o.items():
            if k == "sheetId" and isinstance(v, int):
                acc.add(v)
            else:
                _collect_ids(v, acc)
    elif isinstance(o, list):
        for x in o:
            _collect_ids(x, acc)


def guard(requests):
    acc = set()
    _collect_ids(requests, acc)
    bad = acc & PROTECTED_IDS
    if bad:
        raise SystemExit(f"GUARD ABORT: requests miram abas protegidas {bad}")


# ----------------------------- io -----------------------------

def load_xlsx():
    creds = Credentials.from_service_account_file(SA, scopes=["https://www.googleapis.com/auth/drive.readonly"])
    drv = build("drive", "v3", credentials=creds, cache_discovery=False)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, drv.files().get_media(fileId=SRC, supportsAllDrives=True))
    done = False
    while not done:
        _, done = dl.next_chunk()
    return openpyxl.load_workbook(io.BytesIO(buf.getvalue()), data_only=False)


def sheets_service():
    creds = Credentials.from_service_account_file(SA, scopes=["https://www.googleapis.com/auth/spreadsheets"])
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


# ----------------------------- build -----------------------------

def build_call2_for_tab(sid, wf, native_title, has_chart):
    req = []
    last_content = used_bounds(wf)
    rows = [{"values": [cell_to_celldata(wf.cell(r, c)) for c in range(1, 21)]}
            for r in range(1, last_content + 1)]
    req.append({"updateCells": {"start": {"sheetId": sid, "rowIndex": 0, "columnIndex": 0},
                                "rows": rows, "fields": "userEnteredValue,userEnteredFormat"}})
    # merges painel
    for (r0, r1, c0, c1) in [(1, 2, 13, 20), (7, 8, 13, 20)]:
        req.append({"mergeCells": {"range": {"sheetId": sid, "startRowIndex": r0, "endRowIndex": r1,
                    "startColumnIndex": c0, "endColumnIndex": c1}, "mergeType": "MERGE_ALL"}})
    # larguras + ocultar D
    for letter, px in WIDTHS.items():
        i = col_idx(letter)
        props, fields = {"pixelSize": px}, "pixelSize"
        if letter == "D":
            props["hiddenByUser"] = True
            fields = "pixelSize,hiddenByUser"
        req.append({"updateDimensionProperties": {"range": {"sheetId": sid, "dimension": "COLUMNS",
                    "startIndex": i, "endIndex": i + 1}, "properties": props, "fields": fields}})
    if native_title == "Metas BSB":
        i = col_idx("L")
        req.append({"updateDimensionProperties": {"range": {"sheetId": sid, "dimension": "COLUMNS",
                    "startIndex": i, "endIndex": i + 1}, "properties": {"hiddenByUser": True},
                    "fields": "hiddenByUser"}})
    # altura row1
    req.append({"updateDimensionProperties": {"range": {"sheetId": sid, "dimension": "ROWS",
                "startIndex": 0, "endIndex": 1}, "properties": {"pixelSize": 30}, "fields": "pixelSize"}})
    # CF nativa (transcrita do xlsx)
    idx = 0
    for cr, rules in wf.conditional_formatting._cf_rules.items():
        sqref = str(getattr(cr, "sqref", cr))
        for rng_str in sqref.split():
            gr = a1_range_to_grid_range(rng_str)
            gr["sheetId"] = sid
            for rule in rules:
                if not getattr(rule, "formula", None):
                    continue
                color = cf_color(rule)
                req.append({"addConditionalFormatRule": {"rule": {"ranges": [dict(gr)], "booleanRule": {
                    "condition": {"type": "CUSTOM_FORMULA", "values": [
                        {"userEnteredValue": "=" + commas_to_semicolons(rule.formula[0])}]},
                    "format": {"backgroundColor": color}}}, "index": idx}})
                idx += 1
    # grafico nativo
    if has_chart:
        req.append(chart_request(sid, native_title.replace("Metas ", ""), last_semana_row(wf)))
    return req


def main(dry=False):
    wbf = load_xlsx()
    svc = sheets_service()
    meta = svc.spreadsheets().get(spreadsheetId=DEST, fields="sheets.properties(sheetId,title)").execute()
    existing = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta["sheets"]}

    # CALL 1: deleta abas de metas antigas (se existirem) + cria as 3 novas
    req1 = []
    for native_title, _x, _h, _c in CITIES:
        if native_title in existing:
            sid = existing[native_title]
            if sid in PROTECTED_IDS:
                raise SystemExit(f"GUARD: '{native_title}' colide com aba protegida {sid}")
            req1.append({"deleteSheet": {"sheetId": sid}})
    for native_title, _x, hidden, _c in CITIES:
        props = {"title": native_title,
                 "gridProperties": {"rowCount": 60, "columnCount": 26, "frozenRowCount": 1}}
        if hidden:
            props["hidden"] = True
        req1.append({"addSheet": {"properties": props}})

    if dry:
        print(f"[DRY] CALL1: {len(req1)} requests ({sum('deleteSheet' in r for r in req1)} del, "
              f"{sum('addSheet' in r for r in req1)} add)")
        fake = {t: -(i + 1) for i, (t, *_r) in enumerate(CITIES)}
        total2 = 0
        for native_title, xlsx_title, _h, has_chart in CITIES:
            r2 = build_call2_for_tab(fake[native_title], wbf[xlsx_title], native_title, has_chart)
            kinds = {}
            for r in r2:
                k = next(iter(r))
                kinds[k] = kinds.get(k, 0) + 1
            print(f"[DRY] {native_title}: {len(r2)} requests {kinds} (last_content={used_bounds(wbf[xlsx_title])})")
            total2 += len(r2)
        print(f"[DRY] CALL2 total: {total2} requests. Nada enviado.")
        return

    guard(req1)
    rep = svc.spreadsheets().batchUpdate(spreadsheetId=DEST, body={"requests": req1}).execute()
    ids = {p["addSheet"]["properties"]["title"]: p["addSheet"]["properties"]["sheetId"]
           for p in rep["replies"] if "addSheet" in p}
    print("abas criadas:", ids)

    req2 = []
    for native_title, xlsx_title, _h, has_chart in CITIES:
        req2 += build_call2_for_tab(ids[native_title], wbf[xlsx_title], native_title, has_chart)
    guard(req2)
    svc.spreadsheets().batchUpdate(spreadsheetId=DEST, body={"requests": req2}).execute()
    print(f"BUILD OK: {len(req2)} requests aplicados em {len(ids)} abas.")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    main(dry=args.dry_run)
