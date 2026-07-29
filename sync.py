"""
Sync Ticketsports -> Google Sheets + Leadlovers (multi-etapa)
Puxa inscricoes da API Ticketsports para todas as etapas configuradas,
escreve em abas raw separadas no Google Sheets, e envia novos inscritos
ao Leadlovers para disparo da regua de email.
Roda via GitHub Actions (cron horario) ou manualmente.
"""

import argparse
import http.client
import io
import json
import os
import ssl
import time
import urllib.parse
from datetime import datetime, date, timezone, timedelta

import gspread
from google.oauth2.service_account import Credentials

# ===================================================
# CONFIG TICKETSPORTS
# ===================================================

API_BASE = "api.ticketsports.com.br"
API_VERSION = "/v1.0"
PAGE_LIMIT = 50

# Etapas Corrida Vai Bem. Adicionar nova etapa = nova entrada aqui.
EVENTS = [
    {
        "key": "bsb",
        "id": 86595,
        "label": "Brasília",
        "raw_tab": "raw_inscritos_brasilia",
        "dash_tab": "Brasília",
        "ll_sequence_env": "LL_SEQUENCE_BSB",
        "ll_sent_tab": "Etapa Brasilia",
    },
    {
        "key": "bh",
        "id": 86781,
        "label": "Belo Horizonte",
        "raw_tab": "raw_inscritos_bh",
        "dash_tab": "Belo Horizonte",
        "ll_sequence_env": "LL_SEQUENCE_BH",
        "ll_sent_tab": "Etapa BH",
    },
    {
        "key": "ssa",
        "id": 87008,
        "label": "Salvador",
        "raw_tab": "raw_inscritos_ssa",
        "dash_tab": "Salvador",
        "ll_sequence_env": "LL_SEQUENCE_SSA",
        "ll_sent_tab": "Etapa Salvador",
    },
    {
        "key": "pedalx",
        "id": 87735,
        "label": "Pedal X Road — Brasília",
        "raw_tab": "raw_inscritos_pedalx",
        "dash_tab": "Pedal X Road",
        "ll_sequence_env": "LL_SEQUENCE_PEDALX",
        "ll_sent_tab": "Etapa Pedal X Road",
        "timestamp_cell": "F2",
        "non_blocking": True,
        "expected_modalidades": {"Inscreva-se"},
        "expected_categorias": {"Pedal X"},
    },
    # MTBs. ll_sequence_env aponta para secrets que NAO existem: sem regua de e-mail
    # nessas etapas (sync_to_leadlovers retorna cedo quando a variavel esta vazia).
    # expected_* fica so no builder, que tem a lista completa de categorias etarias;
    # aqui ficam vazios para nao duplicar 38 strings em dois lugares.
    {
        "key": "pedalx_manaus",
        "id": 87732,
        "label": "Pedal X MTB — Manaus",
        "raw_tab": "raw_inscritos_pedalx_manaus",
        "dash_tab": "Pedal X Manaus",
        "ll_sequence_env": "LL_SEQUENCE_PEDALX_MANAUS",
        "ll_sent_tab": "Etapa Pedal X Manaus",
        "timestamp_cell": "F2",
        "non_blocking": True,
        "expected_categorias": {"Kit PedalX - XCO", "Kit PedalX - CATEGORIA PRÓ - 60 km"},
    },
    {
        "key": "pedalx_canastra",
        "id": 87727,
        "label": "Pedal X MTB — Serra da Canastra",
        "raw_tab": "raw_inscritos_pedalx_canastra",
        "dash_tab": "Pedal X Canastra",
        "ll_sequence_env": "LL_SEQUENCE_PEDALX_CANASTRA",
        "ll_sent_tab": "Etapa Pedal X Canastra",
        "timestamp_cell": "F2",
        "non_blocking": True,
        "expected_categorias": {"CATEGORIA SPORT - 30 km ", "CATEGORIA PRÓ - 60 km"},
    },
    # Circuito Santos. UNICO evento de OUTRO organizador (EGP BRASIL, 22.087.202/0001-55).
    # `Order/List` e escopado pelo CNPJ do login, entao a conta padrao devolve 204 vazio aqui
    # — sem erro, indistinguivel de "loja sem venda". Por isso a etapa usa credencial propria
    # (`login_env`/`password_env`), de um usuario que a EGP criou na conta deles em 29/07.
    # Se esse secret sumir, `token_for_event` LEVANTA e o non_blocking segura: melhor a etapa
    # falhar visivelmente do que voltar a 204 silencioso.
    {
        "key": "santos",
        "id": 87817,
        "label": "Circuito Santos",
        "raw_tab": "raw_inscritos_santos",
        "dash_tab": "Circuito Santos",
        "login_env": "TICKET_LOGIN_SANTOS",
        "password_env": "TICKET_PASSWORD_SANTOS",
        "ll_sequence_env": "LL_SEQUENCE_SANTOS",
        "ll_sent_tab": "Etapa Circuito Santos",
        "timestamp_cell": "F2",
        "non_blocking": True,
        "expected_modalidades": {"Corrida 5km", "Caminhada 5km"},
        "expected_categorias": {"KIT ATLETA", "KIT ATLETA - CORTESIAS"},
    },
]

# Credenciais via variáveis de ambiente (GitHub Secrets) ou arquivo local
TICKET_LOGIN = os.environ.get("TICKET_LOGIN", "marketing@brada.social")
TICKET_PASSWORD = os.environ.get("TICKET_PASSWORD", "102030")
SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID", "")
SERVICE_ACCOUNT_JSON = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
SERVICE_ACCOUNT_FILE = os.environ.get(
    "GOOGLE_SERVICE_ACCOUNT_FILE",
    r"C:\Users\bruno\.brada-secrets\sheets-sa.json",
)

HEADER = [
    "N inscricao", "Categoria", "Modalidade", "Sexo", "Status do pedido",
    "Cupom", "Valor", "Data Pedido", "Dispositivo", "Cidade", "Estado", "Camiseta",
    "Inscricao Grupo", "Nome Grupo",
]

# ===================================================
# CONFIG LEADLOVERS
# ===================================================

LL_API_BASE = "llapi.leadlovers.com"
LL_API_TOKEN = os.environ.get("LL_API_TOKEN", "")
LL_MACHINE_CODE = os.environ.get("LL_MACHINE_CODE", "")
# Sequence code e por etapa: cada cidade tem grupo de WhatsApp diferente,
# entao precisa de sequencia propria. Lido via os.environ[ev["ll_sequence_env"]].

# Planilha separada para logs do Leadlovers (nao misturar com dashboard).
# Cada etapa tem aba propria — nome em EVENTS[i]["ll_sent_tab"].
LL_SPREADSHEET_ID = "1aaDYxjcDhhR2lMLejpOW54QVNGQuSOXc8Gj6q5S2KdA"
LL_SENT_HEADER = ["inscricao", "email", "nome", "data_envio"]


# ===================================================
# CONFIG METAS (planilha semanal da Tamyris — meta_corrida_vai_bem)
# ===================================================

# ID do arquivo .xlsx no Drive (mesmo link que o time usa). Gravacao in-place via Drive API.
# `or` (nao o default do get): secret inexistente no Actions vira string vazia, que sombrearia
# o default; assim vazio/ausente cai no ID hardcoded.
METAS_SPREADSHEET_ID = os.environ.get("METAS_SPREADSHEET_ID") or "1t5xEHgT-g6k9wAWspjXKDMssX0rNYhJS"
# Dry-run: imprime o que escreveria, sem tocar a planilha. METAS_DRY_RUN=1 liga.
METAS_DRY_RUN = os.environ.get("METAS_DRY_RUN", "").strip().lower() not in ("", "0", "false", "no")
CAMPAIGN_YEAR = 2026

# UMA aba por cidade (tabela Pagas semanal + bloco por tier + resumo Gratuitas, tudo junto).
METAS_TABS = {
    86595: "Metas [ BSB ]",
    86781: "Metas [ BH ]",
    87008: "Metas [ SSA ]",
}

# Abas que ganham o grafico grande de evolucao. AUTO-CURA: o cron recria o grafico a cada run
# porque o Google Sheets descarta grafico embutido (openpyxl) toda vez que um HUMANO edita o
# .xlsx. O cron preserva; so a edicao humana derruba -> recriar de hora em hora repara em <=1h.
# (CF/formula/SPARKLINE/estilo sobrevivem a edicao humana; so o grafico-objeto morre.)
# BSB (encerrada/oculta) fica de fora.
METAS_CHART_TABS = {86781, 87008}  # BH, SSA
METAS_SELFHEAL_CHART = os.environ.get("METAS_SELFHEAL_CHART", "1").strip().lower() not in ("0", "false", "no")

# Colunas de detalhamento por tier que a automacao controla nas abas Pagas:
# (header na planilha, chave no dict de contagem).
# `Realizado` (E) ja e o total pago, entao NAO ha coluna "Real. Total Pago" (era duplicata).
META_TIER_COLS = [
    ("Real. Básico", "Básico"),
    ("Real. Premium", "Premium"),
    ("Real. Combo", "Combo"),
    ("Real. PCD", "PCD"),
    ("Real. Gratuito", "Gratuito"),
]

# --- Camada NATIVA de metas (abas no proprio Dashboard, gravadas via gspread) ---
# Backend: "native" (abas nativas), "xlsx" (arquivo antigo), "both" (dual-write na migracao).
# Default "both" durante a janela de validacao; depois "native". Ver setup_metas_native.py (build).
# `or "both"` (nao o default do get): a var do Actions vazia/ausente vira "" -> cairia em nenhum
# backend (regressao: para o xlsx E nao roda nativo). Valor invalido (typo) tambem cai em "both".
_mb = (os.environ.get("METAS_BACKEND") or "both").strip().lower()
METAS_BACKEND = _mb if _mb in ("native", "xlsx", "both") else "both"
# Abas nativas (nomes limpos — nativo aceita, ao contrario do .xlsx que proibia colchetes).
# Os 3 pedais e o Circuito Santos sao NATIVO-ONLY: nao entram em METAS_TABS (o .xlsx e da
# Tamyris e so tem corridas) nem em METAS_CHART_TABS (auto-cura e do grafico do .xlsx; o
# nativo e duravel). Build das abas: setup_metas_pedal.py.
METAS_TABS_NATIVE = {
    86595: "Metas BSB",
    86781: "Metas BH",
    87008: "Metas SSA",
    87735: "Metas Pedal Road",
    87732: "Metas Pedal Manaus",
    87727: "Metas Pedal Canastra",
    87817: "Metas Circuito Santos",
}
# Abas cujo build e o setup_metas_pedal.py (as demais vem do setup_metas_native.py). Usado
# so para sugerir o builder certo na mensagem de "aba nao encontrada".
METAS_TABS_BUILDER_PEDAL = {87735, 87732, 87727, 87817}
# Colunas FIXAS (letra) do detalhamento por tier na tabela semanal nativa.
META_TIER_COLS_NATIVE = [("G", "Básico"), ("H", "Premium"), ("I", "Combo"), ("J", "PCD"), ("K", "Gratuito")]


# ===================================================
# RETRY
# ===================================================

def _retry(fn, label, max_tries=3, initial_wait=15):
    """Executa fn, retentando ate max_tries vezes com backoff em caso de falha."""
    wait = initial_wait
    for attempt in range(1, max_tries + 1):
        try:
            return fn()
        except Exception as e:
            if attempt == max_tries:
                raise
            print(f"  [{label}] tentativa {attempt}/{max_tries} falhou: {e}. Aguardando {wait}s...")
            time.sleep(wait)
            wait *= 2


# ===================================================
# API TICKETSPORTS
# ===================================================

def api_request(method, endpoint, headers=None, body=None):
    """Faz request HTTP para a API Ticketsports. Suporta GET com body."""
    def _do():
        ctx = ssl.create_default_context()
        conn = http.client.HTTPSConnection(API_BASE, context=ctx)
        body_str = json.dumps(body) if body else None
        all_headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
        }
        if headers:
            all_headers.update(headers)
        if body_str:
            all_headers["Content-Length"] = str(len(body_str.encode("utf-8")))
        conn.request(method, API_VERSION + endpoint, body=body_str, headers=all_headers)
        resp = conn.getresponse()
        data = resp.read().decode("utf-8")
        conn.close()
        if resp.status == 204 or not data.strip():
            return {}  # sem conteudo (204 ou body vazio — ex: Cortesia com 0 resultados)
        if resp.status != 200:
            raise Exception(f"HTTP {resp.status} em {endpoint}: {data[:300]}")
        return json.loads(data)

    return _retry(_do, endpoint.split("?")[0])


def authenticate(login=None, password=None):
    """Autentica na API e retorna o Bearer token. Sem argumento usa a conta padrao."""
    login = login or TICKET_LOGIN
    password = password or TICKET_PASSWORD

    def _do():
        # quote_plus nos dois campos: senha forte com & = + quebraria o corpo do form.
        body = (f"Login={urllib.parse.quote_plus(login)}"
                f"&Password={urllib.parse.quote_plus(password)}&AccessType=O")
        ctx = ssl.create_default_context()
        conn = http.client.HTTPSConnection(API_BASE, context=ctx)
        conn.request(
            "POST",
            API_VERSION + "/Access",
            body=body,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Accept": "application/json",
            },
        )
        resp = conn.getresponse()
        data = json.loads(resp.read().decode("utf-8"))
        conn.close()
        if not data.get("access_token"):
            raise Exception(f"Falha na autenticação: {data}")
        return data["access_token"]

    token = _retry(_do, "auth")
    print("Autenticado com sucesso.")
    return token


def token_for_event(event, cache):
    """Token da conta que enxerga ESTE evento, reusando por credencial.

    `Order/List` e escopado pelo CNPJ do organizador do login: um evento de outro produtor
    devolve 204 vazio com a conta padrao, sem erro nenhum. Eventos assim declaram
    `login_env`/`password_env` e usam credencial propria.

    Se o evento declara credencial mas o secret nao existe, LEVANTA em vez de cair na conta
    padrao — o fallback silencioso daria 204, que e indistinguivel de "loja sem venda" e
    esconderia um secret faltando por semanas.
    """
    chave = event.get("login_env") or "_default"
    if chave not in cache:
        if chave == "_default":
            cache[chave] = authenticate()
        else:
            login = os.environ.get(event["login_env"], "")
            senha = os.environ.get(event.get("password_env", ""), "")
            if not login or not senha:
                raise RuntimeError(
                    f"{event['label']} exige a credencial {event['login_env']}/"
                    f"{event.get('password_env')} e ela nao esta no ambiente"
                )
            cache[chave] = authenticate(login, senha)
    return cache[chave]


def fetch_all_orders(token, event_id):
    """Busca todos os pedidos pagos de um evento, paginando. Retorna list[dict]."""
    all_participants = []
    page = 1
    total_pages = 1

    while page <= total_pages:
        endpoint = f"/Order/List?page={page}&limit={PAGE_LIMIT}"
        data = api_request(
            "GET",
            endpoint,
            headers={"Authorization": f"Bearer {token}"},
            body={"events": [event_id], "status": ["Pago", "Cortesia"]},
        )

        total_pages = data.get("totalpages", data.get("totalPages", 1))
        orders = data.get("orders", [])

        for order in orders:
            participantes = order.get("participante", [])
            if not isinstance(participantes, list):
                participantes = [participantes]

            for p in participantes:
                camiseta = ""
                produtos = p.get("produtos", [])
                if produtos and produtos[0].get("Camisetas"):
                    camiseta = produtos[0]["Camisetas"]

                cidade = p.get("cidade", "") or order.get("responsavel", {}).get("cidade", "")
                estado = p.get("estado", "") or order.get("responsavel", {}).get("estado", "")

                valor = p.get("valorUnitario", "")
                if isinstance(valor, str):
                    valor = valor.replace(",", ".")

                all_participants.append({
                    "inscricao": p.get("inscricao", ""),
                    "nome": p.get("nome", ""),
                    "email": p.get("email", ""),
                    "celular": p.get("celular", ""),
                    "categoria": p.get("categoria", ""),
                    "modalidade": p.get("modalidade", ""),
                    "sexo": p.get("sexo", ""),
                    "status": order.get("status", ""),
                    "cupom": p.get("tituloCupom", ""),
                    "valor": valor,
                    "dataPedido": order.get("dataPedido", ""),
                    "dispositivo": order.get("tipoDispositivo", ""),
                    "cidade": cidade,
                    "estado": estado,
                    "camiseta": camiseta,
                    "inscricao_grupo": "Sim" if p.get("inscricao_grupo") else "Não",
                    "nome_grupo": p.get("nome_grupo", "") or "",
                })

        print(f"  Página {page}/{total_pages} ({len(orders)} pedidos)")
        page += 1

    return all_participants


def to_sheet_row(p):
    """Converte dict de participante para lista de 14 colunas do Sheet."""
    return [
        p["inscricao"], p["categoria"], p["modalidade"], p["sexo"],
        p["status"], p["cupom"], p["valor"], p["dataPedido"],
        p["dispositivo"], p["cidade"], p["estado"], p["camiseta"],
        p["inscricao_grupo"], p["nome_grupo"],
    ]


# ===================================================
# GOOGLE SHEETS
# ===================================================

def get_credentials(scopes=None):
    """Credenciais da service account (compartilhadas entre gspread e Drive API)."""
    if scopes is None:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
    if SERVICE_ACCOUNT_JSON:
        return Credentials.from_service_account_info(json.loads(SERVICE_ACCOUNT_JSON), scopes=scopes)
    if os.path.exists(SERVICE_ACCOUNT_FILE):
        return Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
    raise Exception(
        "Credenciais Google não encontradas. "
        "Defina GOOGLE_SERVICE_ACCOUNT_JSON ou GOOGLE_SERVICE_ACCOUNT_FILE."
    )


def get_sheets_client():
    """Cria cliente gspread autenticado via service account."""
    return gspread.authorize(get_credentials())


def migrate_legacy_tab(sh):
    """Renomeia raw_inscritos -> raw_inscritos_brasilia (one-shot, idempotente)."""
    try:
        legacy = sh.worksheet("raw_inscritos")
    except gspread.exceptions.WorksheetNotFound:
        return  # Já migrado
    try:
        sh.worksheet("raw_inscritos_brasilia")
        # Destino já existe — apaga legacy pra evitar conflito
        sh.del_worksheet(legacy)
        print("Legacy raw_inscritos removido (raw_inscritos_brasilia ja existia).")
    except gspread.exceptions.WorksheetNotFound:
        legacy.update_title("raw_inscritos_brasilia")
        print("Renomeado: raw_inscritos -> raw_inscritos_brasilia")


def write_raw_tab(sh, rows, raw_tab_name):
    """Sobrescreve uma aba raw com os dados frescos."""
    if not rows:
        raise ValueError(
            f"API retornou zero participantes para {raw_tab_name}; raw preservada."
        )

    try:
        ws = sh.worksheet(raw_tab_name)
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=raw_tab_name, rows=max(1000, len(rows) + 10), cols=len(HEADER))

    ws.clear()
    _retry(lambda: ws.update(values=[HEADER] + rows, range_name="A1"), f"write {raw_tab_name}")
    print(f"  -> {len(rows)} linhas em {raw_tab_name}")


def update_timestamps(sh, events):
    """Escreve timestamp na célula configurada; strings legadas usam C2."""
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    for event in events:
        if isinstance(event, str):
            tab = event
            cell = "C2"
        else:
            tab = event["dash_tab"]
            cell = event.get("timestamp_cell", "C2")
        try:
            ws = sh.worksheet(tab)
            ws.update(values=[[now]], range_name=cell)
        except Exception as e:
            print(f"  Aviso: timestamp {tab}!{cell}: {e}")


# ===================================================
# LEADLOVERS
# ===================================================

def get_ll_sheet(gc):
    """Abre a planilha separada de logs do Leadlovers."""
    return gc.open_by_key(LL_SPREADSHEET_ID)


def get_sent_inscricoes(ll_sh, tab_name):
    """Retorna set de inscricao IDs já enviados ao Leadlovers para a etapa.

    Garante que a aba existe e tem header. Idempotente.
    """
    try:
        ws = ll_sh.worksheet(tab_name)
    except gspread.exceptions.WorksheetNotFound:
        ws = ll_sh.add_worksheet(title=tab_name, rows=5000, cols=len(LL_SENT_HEADER))

    values = ws.col_values(1)  # coluna inscricao
    if not values:
        ws.update(values=[LL_SENT_HEADER], range_name="A1")
        return set()
    return set(str(v) for v in values[1:])  # pula header


def mark_sent_inscricoes(ll_sh, tab_name, new_entries):
    """Appenda novas linhas na aba da etapa."""
    ws = ll_sh.worksheet(tab_name)
    _retry(lambda: ws.append_rows(new_entries), f"log LL {tab_name}")


def push_to_leadlovers(ll_sh, participants, event):
    """Envia novos inscritos (não enviados antes) ao Leadlovers.

    Cada etapa usa sua propria sequencia (link de WhatsApp diferente por cidade).
    """
    event_label = event["label"]

    if not LL_API_TOKEN:
        print("  [LL] LL_API_TOKEN não configurado — pulando sync Leadlovers.")
        return

    sequence_env = event["ll_sequence_env"]
    sequence_code = os.environ.get(sequence_env, "")
    if not sequence_code:
        print(f"  [LL] {sequence_env} não configurado — pulando {event_label}.")
        return

    sent_tab = event["ll_sent_tab"]
    sent = get_sent_inscricoes(ll_sh, sent_tab)
    new_participants = [p for p in participants if str(p["inscricao"]) not in sent]
    print(f"  [LL] {len(new_participants)} novos de {len(participants)} total")

    if not new_participants:
        return

    ctx = ssl.create_default_context()
    successful = []
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    for p in new_participants:
        if not p.get("email"):
            print(f"    - inscricao {p['inscricao']} sem email, pulando")
            continue

        payload = {
            "Email": p["email"],
            "Name": p["nome"],
            "MachineCode": int(LL_MACHINE_CODE),
            "EmailSequenceCode": int(sequence_code),
            "SequenceLevelCode": "1",
            "PhoneNumber": p["celular"],
            "City": p["cidade"],
            "State": p["estado"],
        }
        body_str = json.dumps(payload)
        conn = http.client.HTTPSConnection(LL_API_BASE, context=ctx)
        conn.request(
            "POST",
            f"/webapi/lead?Token={LL_API_TOKEN}",
            body=body_str,
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "Content-Length": str(len(body_str.encode("utf-8"))),
            },
        )
        resp = conn.getresponse()
        resp_body = resp.read().decode("utf-8")
        conn.close()

        if resp.status in (200, 201):
            successful.append([str(p["inscricao"]), p["email"], p["nome"], now])
            print(f"    ✓ {p['email']}")
        else:
            print(f"    ✗ {p['email']} — HTTP {resp.status}: {resp_body[:120]}")

    if successful:
        mark_sent_inscricoes(ll_sh, sent_tab, successful)
        print(f"  [LL] {len(successful)} leads enviados com sucesso.")


# ===================================================
# METAS — preenchimento da planilha semanal da Tamyris
# ===================================================

def parse_valor(x):
    """valorUnitario pode vir como numero, string '99.00'/'99,00', None ou ''."""
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_data_pedido(s):
    """'DD/MM/YYYY HH:MM' (com ano) -> date. Tolera segundos e so data. None se falhar."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _norm_dash(s):
    return str(s).replace("–", "-").replace("—", "-")


def parse_periodo_fim(texto, year=CAMPAIGN_YEAR):
    """'DD/MM - DD/MM' -> date da data FINAL (corte cumulativo). None se falhar."""
    if not texto:
        return None
    parts = [p.strip() for p in _norm_dash(texto).split("-")]
    if len(parts) < 2 or not parts[-1]:
        return None
    chunk = parts[-1].split("/")
    try:
        d, m = int(chunk[0]), int(chunk[1])
        y = int(chunk[2]) if len(chunk) > 2 and chunk[2] else year
        return date(y, m, d)
    except (ValueError, IndexError):
        return None


def parse_inicio(texto, year=CAMPAIGN_YEAR):
    """'DD/MM' (ou 'DD/MM/AAAA') -> date. None se falhar."""
    if not texto:
        return None
    chunk = str(texto).strip().split("/")
    try:
        d, m = int(chunk[0]), int(chunk[1])
        y = int(chunk[2]) if len(chunk) > 2 and chunk[2] else year
        return date(y, m, d)
    except (ValueError, IndexError):
        return None


def _today_brt():
    """Data de 'hoje' em horario de Brasilia (UTC-3). O runner do GitHub Actions roda em UTC."""
    return datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=-3))).date()


def _semana_futura(periodo_txt, hoje):
    """True se a semana ainda nao comecou (data de INICIO do periodo 'DD/MM - DD/MM' > hoje)."""
    inicio = parse_inicio(_norm_dash(periodo_txt).split("-")[0].strip())
    return inicio is not None and inicio > hoje


def is_free(p):
    """Gratis: valor 0 OU status Cortesia (cobre perna gratis de combo e cortesias)."""
    return parse_valor(p.get("valor")) < 0.01 or (p.get("status") or "") == "Cortesia"


def is_pcd(p):
    return "PCD" in (p.get("categoria") or "").upper()


def is_combo(p):
    return "COMBO" in (p.get("categoria") or "").upper()


def base_tier(p):
    """Tier base de um inscrito PAGO: Premium se a categoria contem PREMIUM, senao Basico."""
    return "Premium" if "PREMIUM" in (p.get("categoria") or "").upper() else "Básico"


def tier_counts_cumulative(participants, end_date):
    """Contagem cumulativa (dataPedido <= end_date) por tier.

    Total Pago = Basico + Premium (valor>0, nao-cortesia). Combo e PCD sao
    recortes informativos (subconjuntos), nao somam no total.
    """
    c = {"Básico": 0, "Premium": 0, "Combo": 0, "PCD": 0,
         "Gratuito": 0, "Total Pago": 0, "_ignorados": 0}
    for p in participants:
        d = parse_data_pedido(p.get("dataPedido"))
        if d is None:
            c["_ignorados"] += 1
            continue
        if d > end_date:
            continue
        if is_free(p):
            c["Gratuito"] += 1
            continue
        c["Total Pago"] += 1
        c[base_tier(p)] += 1
        if is_combo(p):
            c["Combo"] += 1
        if is_pcd(p):
            c["PCD"] += 1
    return c


def gratuito_count_since(participants, inicio_date):
    """Conta gratuitos (valor 0 ou Cortesia) com dataPedido >= inicio_date (todos se None)."""
    n = 0
    for p in participants:
        d = parse_data_pedido(p.get("dataPedido"))
        if d is None:
            continue
        if inicio_date and d < inicio_date:
            continue
        if is_free(p):
            n += 1
    return n


def _norm_header(s):
    return " ".join(str(s).strip().split()).casefold()


def _to_int(x):
    """Le um inteiro de uma celula (numero ou texto, tolera separador de milhar)."""
    if isinstance(x, (int, float)):
        return int(round(x))
    s = "".join(ch for ch in str(x) if ch.isdigit())
    return int(s) if s else None


# --- Escrita IN-PLACE no proprio .xlsx (Drive API + openpyxl, preserva o mesmo link) ---
# A Sheets API nao escreve em .xlsx. O robo baixa o arquivo fresco, edita SO as celulas
# de Realizado/Gap/Real.* com openpyxl (preserva todo o resto) e sobe nova versao via Drive
# (mesmo fileId = mesmo link). openpyxl mantem valores, estilos e larguras (testado).

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _norm_sheet(s):
    """Nome de aba normalizado: remove colchetes (proibidos em .xlsx), colapsa espacos, casefold."""
    s = str(s).replace("[", " ").replace("]", " ")
    return " ".join(s.split()).casefold()


def _find_ws(wb, logical_name):
    target = _norm_sheet(logical_name)
    for ws in wb.worksheets:
        if _norm_sheet(ws.title) == target:
            return ws
    return None


def _header_map_xlsx(ws):
    return {_norm_header(ws.cell(1, c).value): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(1, c).value not in (None, "")}


def _ensure_cols(ws, names):
    """Garante headers na linha 1 (anexa a direita os que faltam, idempotente). Retorna o mapa."""
    hmap = _header_map_xlsx(ws)
    nxt = ws.max_column + 1
    for nm in names:
        if _norm_header(nm) not in hmap:
            ws.cell(1, nxt).value = nm
            hmap[_norm_header(nm)] = nxt
            nxt += 1
    return hmap


def write_metas_pagas_xlsx(ws, participants, label, hoje=None):
    """Escreve Realizado (valor SEMANAL = quantos entraram naquela semana), Gap (=Meta-Realizado)
    e colunas Real.* por semana. So toca essas celulas.

    Semanas cujo inicio > hoje (BRT) ficam em branco (preenchem quando a semana chega).
    """
    hoje = hoje or _today_brt()
    hmap = _header_map_xlsx(ws)
    col_semana = hmap.get(_norm_header("Semana"))
    col_periodo = hmap.get(_norm_header("Período"))
    col_real = hmap.get(_norm_header("Realizado"))
    col_meta = hmap.get(_norm_header("Meta")) or hmap.get(_norm_header("Meta Vendas Pagas"))
    if not (col_semana and col_periodo and col_real):
        print(f"  [METAS] {label}: faltam cabecalhos Semana/Periodo/Realizado — pulando aba")
        return 0
    # garante colunas de tier + Gap a direita (idempotente)
    hmap = _ensure_cols(ws, [n for n, _ in META_TIER_COLS] + ["Gap"])
    col_gap = hmap.get(_norm_header("Gap"))
    tier_idx = {name: hmap.get(_norm_header(name)) for name, _ in META_TIER_COLS}

    n = 0
    seen = {}
    ignorados = 0
    prev_cum = {}  # acumulado da semana anterior, p/ derivar o valor SEMANAL (cum - prev_cum)
    for r in range(2, ws.max_row + 1):
        semana = ws.cell(r, col_semana).value
        if semana is None or not str(semana).strip():
            continue
        if not str(semana).strip().lower().startswith("semana"):
            continue  # ignora linhas fora da tabela semanal (ex: bloco de metas por tier)
        periodo_txt = str(ws.cell(r, col_periodo).value or "").strip()
        fim = parse_periodo_fim(periodo_txt)
        if fim is None:
            print(f"  [METAS] {label} {semana}: Periodo '{periodo_txt}' nao parseavel — pulando linha")
            continue
        if periodo_txt in seen:
            print(f"  [METAS] {label} {semana}: Periodo duplicado '{periodo_txt}' (= {seen[periodo_txt]}) — revisar datas")
        seen[periodo_txt] = semana
        if _semana_futura(periodo_txt, hoje):
            # semana ainda nao comecou: limpar (None = branco real, p/ CF e grafico ignorarem)
            ws.cell(r, col_real).value = None
            if col_gap:
                ws.cell(r, col_gap).value = None
            for _name, _key in META_TIER_COLS:
                ci = tier_idx.get(_name)
                if ci:
                    ws.cell(r, ci).value = None
            if METAS_DRY_RUN:
                print(f"  [METAS DRY] {label} {semana}: FUTURA ({periodo_txt}) -> branco")
            n += 1
            continue
        cum = tier_counts_cumulative(participants, fim)
        ignorados = max(ignorados, cum["_ignorados"])
        # valor SEMANAL = quantos entraram NAQUELA semana = cumulativo - cumulativo da anterior
        weekly = {k: cum.get(k, 0) - prev_cum.get(k, 0)
                  for k in ("Básico", "Premium", "Combo", "PCD", "Gratuito", "Total Pago")}
        prev_cum = cum
        ws.cell(r, col_real).value = weekly["Total Pago"]
        if col_gap and col_meta:
            # Gap = Meta - Realizado (ambos da semana), como FORMULA pro Google avaliar.
            from openpyxl.utils import get_column_letter
            ws.cell(r, col_gap).value = f"={get_column_letter(col_meta)}{r}-{get_column_letter(col_real)}{r}"
        for name, key in META_TIER_COLS:
            ci = tier_idx.get(name)
            if ci:
                ws.cell(r, ci).value = weekly[key]
        if METAS_DRY_RUN:
            print(f"  [METAS DRY] {label} {semana}: semana={weekly['Total Pago']} (cum={cum['Total Pago']})")
        n += 1
    if ignorados:
        print(f"  [METAS] {label}: {ignorados} inscritos com dataPedido nao parseavel (ignorados)")
    print(f"  [METAS] {label}: {n} semanas processadas")
    return n


def write_metas_gratuitas_xlsx(ws, participants, label):
    """Escreve o Realizado das gratuitas na secao 'Inicio Monitoramento' (em QUALQUER lugar da aba).

    Funciona tanto na aba consolidada (secao mais embaixo) quanto numa aba so de gratuitas (linha 1).
    """
    hdr_row = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if _norm_header(ws.cell(r, c).value) == _norm_header("Início Monitoramento"):
                hdr_row = r
                break
        if hdr_row:
            break
    if hdr_row is None:
        print(f"  [METAS] {label}: secao Gratuitas nao encontrada — pulando")
        return 0
    cols = {_norm_header(ws.cell(hdr_row, c).value): c
            for c in range(1, ws.max_column + 1) if ws.cell(hdr_row, c).value not in (None, "")}
    col_inicio = cols.get(_norm_header("Início Monitoramento"))
    col_meta = cols.get(_norm_header("Meta Gratuitas")) or cols.get(_norm_header("Meta"))
    col_real = cols.get(_norm_header("Realizado"))
    col_gap = cols.get(_norm_header("Gap"))
    drow = hdr_row + 1  # linha de dados, logo abaixo do cabecalho da secao
    inicio = parse_inicio(ws.cell(drow, col_inicio).value) if col_inicio else None
    g = gratuito_count_since(participants, inicio)
    if col_real:
        ws.cell(drow, col_real).value = g
    if col_gap and col_meta and col_real:
        from openpyxl.utils import get_column_letter
        ws.cell(drow, col_gap).value = f"={get_column_letter(col_meta)}{drow}-{get_column_letter(col_real)}{drow}"
    print(f"  [METAS] {label} gratuitas: Realizado={g} (desde {inicio})")
    return 1


# ===================================================
# METAS NATIVAS (abas no proprio Dashboard, via gspread) — substitui o .xlsx
# ===================================================
# Mesma logica de contagem das funcoes xlsx acima (reusada 1:1); muda so a ESCRITA: em vez de
# baixar/editar/subir o .xlsx (openpyxl+Drive), escreve direto as celulas das abas nativas. O
# grafico nativo e duravel -> NAO precisa de auto-cura. Toca SO E/F/G-K por semana + Realizado/Gap
# das gratuitas; nunca A/B/C (Meta/Periodo/Semana, do humano) nem o painel.
# LOCALE (provado por probe): formulas sem separador (`=C{r}-E{r}`) sao imunes; se precisar de
# multi-arg no futuro, usar `;` (pt_BR), nunca `,`. USER_ENTERED parseia no locale da planilha.

def _a1col(n):
    """Indice 1-based -> letra de coluna (1->A, 27->AA)."""
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def _find_gratuitas_native(grid):
    """Acha a secao Gratuitas em QUALQUER lugar: (linha_1based do header, {header_norm: col_1based})."""
    target = _norm_header("Início Monitoramento")
    for i, row in enumerate(grid):
        for j, val in enumerate(row):
            if _norm_header(val) == target:
                cols = {_norm_header(v): k + 1 for k, v in enumerate(row) if str(v).strip()}
                return i + 1, cols
    return None, None


def _build_metas_writes(grid, participants, label, hoje):
    """PURO (sem rede). Recebe o grid (list-of-lists) de uma aba nativa de metas + participantes;
    devolve (updates, clears, msgs) p/ a Sheets API:
      updates = [{"range": "E5:K5", "values": [[Realizado, "=C5-E5", Bas, Prem, Combo, PCD, Grat]]}, ...]
      clears  = ["E7:K7", ...]  (semanas futuras -> branco real)
    Espelha write_metas_pagas_xlsx + write_metas_gratuitas_xlsx (mesma contagem cumulativa-menos-anterior,
    mesmo fill-as-time, mesma deteccao de duplicata)."""
    updates, clears, msgs = [], [], []
    seen, prev_cum, ignorados = {}, {}, 0
    # tabela semanal (linhas onde col A comeca com "Semana")
    for i, row in enumerate(grid):
        r = i + 1
        semana = (row[0] if len(row) > 0 else "").strip()
        if not semana.lower().startswith("semana") or semana.lower() == "semana":
            continue   # pula nao-semanas E o proprio header "Semana" (real e "Semana 0".."Semana N")
        periodo_txt = (row[1] if len(row) > 1 else "").strip()
        fim = parse_periodo_fim(periodo_txt)
        if fim is None:
            msgs.append(f"  [METAS] {label} {semana}: Período '{periodo_txt}' não parseável — pulando")
            continue
        if periodo_txt in seen:
            msgs.append(f"  [METAS] {label} {semana}: Período duplicado '{periodo_txt}' (= {seen[periodo_txt]}) — revisar datas")
        seen[periodo_txt] = semana
        if _semana_futura(periodo_txt, hoje):
            clears.append(f"E{r}:K{r}")   # futura: branco real (None), p/ CF e grafico ignorarem
            continue
        cum = tier_counts_cumulative(participants, fim)
        ignorados = max(ignorados, cum["_ignorados"])
        weekly = {k: cum.get(k, 0) - prev_cum.get(k, 0)
                  for k in ("Básico", "Premium", "Combo", "PCD", "Gratuito", "Total Pago")}
        prev_cum = cum
        vals = [weekly["Total Pago"], f"=C{r}-E{r}",
                weekly["Básico"], weekly["Premium"], weekly["Combo"], weekly["PCD"], weekly["Gratuito"]]
        updates.append({"range": f"E{r}:K{r}", "values": [vals]})
    if ignorados:
        msgs.append(f"  [METAS] {label}: {ignorados} inscritos com dataPedido não parseável (ignorados)")
    # gratuitas (header-relativo; Realizado e col C aqui, nao E)
    hdr_row, cols = _find_gratuitas_native(grid)
    if hdr_row is None:
        msgs.append(f"  [METAS] {label}: seção Gratuitas não encontrada")
    else:
        c_inicio = cols.get(_norm_header("Início Monitoramento"))
        c_meta = cols.get(_norm_header("Meta Gratuitas")) or cols.get(_norm_header("Meta"))
        c_real = cols.get(_norm_header("Realizado"))
        c_gap = cols.get(_norm_header("Gap"))
        drow = hdr_row + 1
        inicio_raw = None
        if c_inicio and drow - 1 < len(grid) and c_inicio - 1 < len(grid[drow - 1]):
            inicio_raw = grid[drow - 1][c_inicio - 1]
        inicio = parse_inicio(inicio_raw)
        g = gratuito_count_since(participants, inicio)
        if c_real:
            updates.append({"range": f"{_a1col(c_real)}{drow}", "values": [[g]]})
        if c_gap and c_meta and c_real:
            updates.append({"range": f"{_a1col(c_gap)}{drow}",
                            "values": [[f"={_a1col(c_meta)}{drow}-{_a1col(c_real)}{drow}"]]})
        msgs.append(f"  [METAS] {label} gratuitas: Realizado={g} (desde {inicio})")
    return updates, clears, msgs


def write_metas_native(sh, participants_por_cidade):
    """Atualiza as abas NATIVAS de metas (BSB/BH/SSA) no proprio Sheet do dashboard. Isolada por aba
    (uma falha nao derruba as outras nem o sync principal). NUNCA cria aba nem grafico (isso e do
    build setup_metas_native.py); so escreve as celulas computadas."""
    hoje = _today_brt()
    for event_id, tab_name in METAS_TABS_NATIVE.items():
        # Etapa AUSENTE do dict != etapa com zero inscritos. Os pedais e o Santos sao
        # `non_blocking`: quando a API falha, main() da `continue` e nunca grava a chave. Sem
        # esta guarda, o `.get(..., [])` devolveria lista vazia e escreveria 0 em todas as
        # semanas passadas, apagando numero bom ate o run seguinte. Lista vazia com a chave
        # presente (loja aberta que ainda nao vendeu) continua escrevendo 0, que e o correto.
        # No Santos esta guarda e o estado PERMANENTE ate o acesso ao evento 87817 chegar.
        if event_id not in participants_por_cidade:
            print(f"  [METAS] {tab_name}: etapa nao sincronizada neste run — pulando (nao zera a aba)")
            continue
        participants = participants_por_cidade[event_id]
        try:
            ws = sh.worksheet(tab_name)
        except gspread.exceptions.WorksheetNotFound:
            builder = ("setup_metas_pedal.py" if event_id in METAS_TABS_BUILDER_PEDAL
                       else "setup_metas_native.py")
            print(f"  [METAS] aba nativa '{tab_name}' não encontrada (rodar {builder}?) — pulando")
            continue
        try:
            grid = _retry(lambda: ws.get_values(), f"metas read {tab_name}")
            hdr = grid[0] if grid else []

            def _h(idx):
                return _norm_header(hdr[idx]) if idx < len(hdr) else ""
            if not (_h(0) == _norm_header("Semana") and _h(4) == _norm_header("Realizado")
                    and _h(5) == _norm_header("Gap")):
                print(f"  [METAS] {tab_name}: header da linha 1 inesperado — pulando aba (schema guard)")
                continue
            updates, clears, msgs = _build_metas_writes(grid, participants, tab_name, hoje)
            for m in msgs:
                print(m)
            if METAS_DRY_RUN:
                print(f"  [METAS DRY] {tab_name}: {len(updates)} writes, {len(clears)} clears (nada enviado)")
                continue
            if clears:
                _retry(lambda: ws.batch_clear(clears), f"metas clear {tab_name}")
            if updates:
                _retry(lambda: ws.batch_update(updates, value_input_option="USER_ENTERED"),
                       f"metas write {tab_name}")
            print(f"  [METAS] {tab_name}: {len(updates)} writes, {len(clears)} clears (nativo)")
        except Exception as e:
            print(f"  [METAS] erro nativo '{tab_name}': {e}")


def ensure_gratuitas_ssa(wb):
    """Cria a aba Gratuitas SSA se nao existir (.xlsx nao aceita colchetes no nome da aba)."""
    if _find_ws(wb, "Metas Gratuitas [ SSA ]") is not None:
        return
    ws = wb.create_sheet(title="Metas Gratuitas  SSA")
    ws.append(["Início Monitoramento", "Meta Gratuitas", "Observação", "Realizado", "Gap"])
    ws.append(["", 300, "Monitorar distribuição e engajamento", "", ""])
    print("  [METAS] aba 'Metas Gratuitas  SSA' criada (Meta 300 provisoria)")


def _last_semana_row(ws):
    """Ultima linha cujo col-A comeca com 'Semana' (limite da tabela semanal)."""
    last = 1
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).strip().lower().startswith("semana"):
            last = r
    return last


def _metas_layout(last):
    """Posicoes DINAMICAS dos blocos abaixo da tabela semanal, pelo nº de semanas (`last` = ultima
    linha de Semana). Assim adicionar semanas nao colide com tier/gratuitas/grafico. Compartilhado
    entre o setup (que monta os blocos) e o cron (que recria o grafico) p/ ambos concordarem.
      tier_row .. tier_row+3 (titulo + Basico/Premium/Total) | grat_row .. grat_row+2 | chart em A{...}
    """
    tier_row = last + 2
    grat_row = last + 7
    chart_anchor = f"A{last + 11}"
    return tier_row, grat_row, chart_anchor


def add_evolucao_chart(ws, last, cidade, anchor="A20"):
    """LineChart Realizado (E) vs Meta (C) por Semana (A), recriado do zero (refs atuais).

    Ancora 'A20' (full-width abaixo da tabela/gratuitas) p/ NAO colidir com o painel (cols N:T).
    Eixo comeca na Semana 1 (linha 3), pulando a 'Semana 0' (carga pre-lancamento) que senao
    vira um 1o ponto gigante. AUTO-CURA: ws._charts=[] antes -> 1 chart por run (dedup); chamado
    pelo cron a cada hora porque a edicao humana no Google derruba o grafico embutido.
    """
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    ws._charts = []
    if last < 3:
        return  # sem semanas suficientes p/ um grafico
    chart = LineChart()
    chart.title = f"Evolução semanal {cidade}"
    chart.style = 2
    chart.height = 8
    chart.width = 18
    chart.y_axis.title = "Inscricoes"
    chart.x_axis.title = "Semana"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    cats = Reference(ws, min_col=1, min_row=3, max_row=last)
    real = Reference(ws, min_col=5, min_row=3, max_row=last)
    meta = Reference(ws, min_col=3, min_row=3, max_row=last)
    chart.add_data(real, titles_from_data=False)
    chart.add_data(meta, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].tx = SeriesLabel(v="Realizado")
    chart.series[1].tx = SeriesLabel(v="Meta")
    chart.series[0].graphicalProperties.line.solidFill = "C55A11"
    chart.series[0].graphicalProperties.line.width = 28000
    chart.series[1].graphicalProperties.line.solidFill = "999999"
    chart.series[1].graphicalProperties.line.dashStyle = "dash"
    ws.add_chart(chart, anchor)


def _drive_service():
    from googleapiclient.discovery import build
    creds = get_credentials(["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _download_xlsx(drive, file_id):
    from googleapiclient.http import MediaIoBaseDownload

    def _do():
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(buf, drive.files().get_media(fileId=file_id, supportsAllDrives=True))
        done = False
        while not done:
            _, done = dl.next_chunk()
        return buf.getvalue()

    return _retry(_do, "metas download")


def _upload_xlsx(drive, file_id, data):
    from googleapiclient.http import MediaIoBaseUpload

    def _do():
        media = MediaIoBaseUpload(io.BytesIO(data), mimetype=XLSX_MIME, resumable=False)
        return drive.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()

    return _retry(_do, "metas upload")


def sync_metas(participants_por_cidade):
    """Atualiza a planilha de metas (.xlsx in-place via Drive). Isolada: nunca derruba o sync."""
    import openpyxl
    if not METAS_SPREADSHEET_ID:
        print("  [METAS] METAS_SPREADSHEET_ID nao configurado — pulando metas.")
        return
    if METAS_DRY_RUN:
        print("  [METAS] *** DRY-RUN: nada sera enviado ao Drive ***")
    drive = _drive_service()
    try:
        data = _download_xlsx(drive, METAS_SPREADSHEET_ID)
    except Exception as e:
        print(f"  [METAS] nao consegui baixar a planilha: {e}")
        print("  [METAS] confira: Drive API ativa, arquivo compartilhado com a SA, METAS_SPREADSHEET_ID correto.")
        return
    wb = openpyxl.load_workbook(io.BytesIO(data))
    total = 0
    for event_id, tab_name in METAS_TABS.items():
        parts = participants_por_cidade.get(event_id, [])
        ws = _find_ws(wb, tab_name)
        if ws is None:
            print(f"  [METAS] aba '{tab_name}' nao encontrada — pulando")
            continue
        try:
            total += write_metas_pagas_xlsx(ws, parts, tab_name)
        except Exception as e:
            print(f"  [METAS] erro pagas '{tab_name}': {e}")
        try:
            total += write_metas_gratuitas_xlsx(ws, parts, tab_name)
        except Exception as e:
            print(f"  [METAS] erro gratuitas '{tab_name}': {e}")
        # Auto-cura do grafico grande: recria a cada run (a edicao humana no Google o derruba).
        if event_id in METAS_CHART_TABS and METAS_SELFHEAL_CHART and not METAS_DRY_RUN:
            try:
                sigla = tab_name.replace("Metas", "").replace("[", "").replace("]", "").strip()
                lsr = _last_semana_row(ws)
                _, _, anchor = _metas_layout(lsr)
                add_evolucao_chart(ws, lsr, sigla, anchor=anchor)
            except Exception as e:
                print(f"  [METAS] reseed chart '{tab_name}': {e}")
    if METAS_DRY_RUN:
        print(f"  [METAS] DRY-RUN: {total} blocos calculados, planilha NAO enviada.")
        return
    out = io.BytesIO()
    wb.save(out)
    _upload_xlsx(drive, METAS_SPREADSHEET_ID, out.getvalue())
    print(f"  [METAS] planilha atualizada in-place ({total} blocos).")


# ===================================================
# MAIN
# ===================================================

def check_event_schema(participants, event):
    """Alerta quando surge modalidade/categoria FORA do contrato conhecido.

    So o excedente e alertado. Contrato sem venda ainda ("missing") e o estado normal de
    loja recem-aberta — um MTB abre com dezenas de categorias etarias e quase nenhuma
    vendida —, entao alertar nisso viraria ruido de hora em hora. O sinal acionavel e o
    inverso: valor novo que o dashboard nao tem linha para contar, logo venda invisivel.

    Compara com os dois lados normalizados: a TicketSports entrega titulos com espaco
    sobrando (ex.: "CATEGORIA SPORT - 30 km ") e o contrato copia esse literal.
    """
    checks = (
        ("modalidades", "modalidade", event.get("expected_modalidades")),
        ("categorias", "categoria", event.get("expected_categorias")),
    )
    warnings = []
    for label, field, expected in checks:
        if not expected:
            continue
        observed = {str(p.get(field, "")).strip() for p in participants if p.get(field)}
        expected = {str(x).strip() for x in expected}
        unexpected = observed - expected
        if unexpected:
            warnings.append((label, {"missing": expected - observed, "unexpected": unexpected}))
            print(
                f"  [SCHEMA WARNING] {event['label']}: {label} fora do contrato; "
                f"inesperado={sorted(unexpected)!r}, esperado={sorted(expected)!r}"
            )
    return warnings


def sync_event(token, sh, event, ll_sh=None, send_leads=True):
    """Sincroniza um evento; o caller decide isolamento e efeitos auxiliares."""
    print(f"\n[{event['label']}] event_id={event['id']}")
    participants = fetch_all_orders(token, event["id"])
    print(f"  Total: {len(participants)} inscritos")
    check_event_schema(participants, event)
    rows = [to_sheet_row(p) for p in participants]
    write_raw_tab(sh, rows, event["raw_tab"])
    if send_leads:
        if ll_sh is None:
            raise ValueError("ll_sh é obrigatório quando send_leads=True")
        push_to_leadlovers(ll_sh, participants, event)
    return participants


def _parse_args(argv=None):
    parser = argparse.ArgumentParser(description="Sync TicketSports -> Sheets")
    parser.add_argument(
        "--event",
        choices=[event["key"] for event in EVENTS],
        help="Seleciona um único evento; permitido somente com --raw-only.",
    )
    parser.add_argument(
        "--raw-only",
        action="store_true",
        help="Atualiza somente a raw do evento selecionado, sem LL/timestamps/metas.",
    )
    args = parser.parse_args(argv)
    if args.raw_only and not args.event:
        parser.error("--raw-only exige --event")
    if args.event and not args.raw_only:
        parser.error("--event só pode ser usado com --raw-only")
    return args


def _open_dashboard(gc, require_id=False):
    if SPREADSHEET_ID:
        return _retry(lambda: gc.open_by_key(SPREADSHEET_ID), "abrir dashboard")
    if require_id:
        raise ValueError(
            "SPREADSHEET_ID é obrigatório em --raw-only para impedir destino ambíguo."
        )
    return _retry(lambda: gc.open("Dashboard Inscrições - Vai Bem"), "abrir dashboard")


def main(argv=None):
    args = _parse_args(argv)
    print(f"=== Sync Ticketsports -> Sheets + Leadlovers ({datetime.now()}) ===")

    tokens = {}
    gc = get_sheets_client()
    sh = _open_dashboard(gc, require_id=args.raw_only)

    if args.raw_only:
        event = next(event for event in EVENTS if event["key"] == args.event)
        participants = sync_event(token_for_event(event, tokens), sh, event, send_leads=False)
        print(f"\n=== Raw-only concluído: {len(participants)} inscritos em {event['raw_tab']} ===")
        return

    migrate_legacy_tab(sh)
    ll_sh = _retry(lambda: get_ll_sheet(gc), "abrir planilha LL")

    total_inscritos = 0
    participants_por_cidade = {}
    successful_events = []
    for event in EVENTS:
        try:
            participants = sync_event(token_for_event(event, tokens), sh, event,
                                      ll_sh=ll_sh, send_leads=True)
        except Exception as exc:
            if event.get("non_blocking"):
                print(f"  [NON-BLOCKING] {event['label']} falhou; demais etapas seguem: {exc}")
                continue
            raise
        participants_por_cidade[event["id"]] = participants
        successful_events.append(event)
        total_inscritos += len(participants)

    update_timestamps(sh, successful_events)

    # Metas: ultima etapa, isolada — nunca pode derrubar o sync de raw/Leadlovers.
    # Backend via METAS_BACKEND: "both" (migracao), "native" (so abas nativas), "xlsx" (so arquivo).
    try:
        if METAS_BACKEND in ("xlsx", "both"):
            print(f"\n[METAS] (.xlsx) atualizando planilha de metas... [backend={METAS_BACKEND}]")
            try:
                sync_metas(participants_por_cidade)
            except Exception as e:
                print(f"  [METAS] backend xlsx falhou (seguiu): {e}")
        if METAS_BACKEND in ("native", "both"):
            print(f"\n[METAS] (nativo) atualizando abas nativas de metas... [backend={METAS_BACKEND}]")
            try:
                write_metas_native(sh, participants_por_cidade)
            except Exception as e:
                print(f"  [METAS] backend nativo falhou (seguiu): {e}")
    except Exception as e:
        print(f"  [METAS] etapa de metas falhou (sync principal seguiu OK): {e}")

    print(f"\n=== Concluído: {total_inscritos} inscritos em {len(successful_events)} etapas ===")


if __name__ == "__main__":
    main()
